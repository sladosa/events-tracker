"""
Events Tracker - Unified Excel Events I/O Module V2.5.6
========================================================
Created: 2025-01-07 17:00 UTC
Last Modified: 2025-01-13 15:30 UTC
Python: 3.11
Version: 2.5.6

Description:
Unified Excel Export/Import for events with enhanced formatting and LEGEND-BASED import.

BUGFIXES in V2.5.6:
- 🐛 FIXED: Export merging now works correctly!
  - Group by (timestamp, comment) instead of just timestamp
  - Prevents merging unrelated events with same timestamp
  - Example: "novi 2" + "test 2 - corr" no longer merge (different comments)
  - Now properly merges only parent-child events from SAME import row ✅

MAJOR UX IMPROVEMENTS in V2.5.5:
- 🎯 EXPORT: Session-based merging for cleaner Excel output
  - Events with same timestamp in parent-child hierarchy merge into ONE row
  - Example: Cardio + Running (same timestamp) → exports as Running with ALL attributes
  - Reduces Excel rows significantly (from 11 to 6 rows in typical case)
  - Only leaf (deepest) events exported, but with merged parent attributes
  - Independent events (same level) still export separately ✅
- 📊 EXPORT: Respects UI sort order (newest/oldest first)
  - Export now matches what user sees in Show Events UI
  - Consistent UX between view and export ✅

CRITICAL FIX in V2.5.4:
- 🐛 FIXED: UPDATE path now supports multi-level event creation!
  - When updating event and populating parent category attributes → creates parent events
  - Example: Update "Cardio > Running" + populate Cardio attrs → creates NEW Cardio event
  - Parent event shares SAME session_start timestamp as child event
  - Enables full multi-level workflow for both CREATE and UPDATE
  - Previously: UPDATE only modified existing event, ignored parent attributes
  - Now: UPDATE checks hierarchy, creates parent events if attributes populated ✅

CRITICAL FIXES in V2.5.3:
- 🐛 FIXED: session_start default (09:00) now properly applied!
  - Empty or '09:00' → creates timestamp with 09:00 (not NULL)
  - CREATE and UPDATE paths both fixed
  - All events now have valid session_start timestamps ✅
- 🌳 FIXED: Multi-level event creation from single Excel row!
  - One row can create MULTIPLE events (one per hierarchy level)
  - Example: Row with Cardio + Running attrs → creates 2 events
  - All events from same row share SAME session_start timestamp
  - Only creates event if BAREM 1 attribute for that level is populated
  - Only populated attributes saved (empty attributes skipped)
  - Enables true session/activity grouping by timestamp ✅

CRITICAL FIXES in V2.5.2:
- 🐛 FIXED: Parent attributes now ACTUALLY BLUE (was still yellow!)
  - load_categories_dict() missing parent_category_id field
  - Color logic couldn't walk hierarchy without it
  - NOW WORKS! ✅
- 🐛 FIXED: Hierarchical sort now PERFECT (matches Structure Viewer)
  - New algorithm: builds sort_path by walking parent chain
  - Example: "0001/0001/0002" = Strength > Legs > Squats
  - String sort maintains perfect tree order
  - All Categories export: Strength tree → Cardio tree → etc.

BUGFIXES in V2.5.1:
- 🐛 FIXED: Filter logic - only selected category branch exported (no mixing branches)
  - Added get_all_descendant_category_ids() to expand parent → all children
  - When "Cardio" selected, only Cardio tree exported (not Strength)
- 🐛 FIXED: Parent attributes now properly BLUE (not orange) for child events
  - Improved color logic - walks up hierarchy to check relevance
  - Running events now show Cardio attrs as BLUE ✅
- 🐛 FIXED: Attribute sorting now truly hierarchical
  - Parents always before children in same branch

NEW in V2.5.0 - MAJOR RESTRUCTURE:
- 🎯 HIERARCHICAL ATTRIBUTE SORTING: Attributes sorted by category level → sort_order
  - Parent categories first (e.g., Cardio before Cardio > Running)
  - Sub-categories follow their parents in tree order
  - Matches Interactive Structure Viewer display order
- 🌳 PARENT ATTRIBUTES INCLUDED: Export includes all parent category attributes
  - Running events now show Cardio parent attributes (total_duration, avg_hr, etc.)
  - All parent attributes are EDITABLE (blue) for child category events
  - Enables full data entry for inherited attributes
- 🏷️ UNIQUE ATTRIBUTE IDENTIFICATION: Handles duplicate attribute names correctly
  - Uses (category_path, attr_name) as unique key internally
  - LEGEND shows full Category_Path for each attribute
  - EVENT DATA headers show "attr_name (Category)" format
  - No more missing attributes due to name collisions!
- ⏰ TIME COLUMN ADDED: New session_start column for event time
  - Column after event_date, before comment
  - Default value: 09:00 if not specified
  - Format: HH:MM (24-hour)
  - Stored as timestamp in database
- 🎨 IMPROVED LEGEND CLARITY: Category_Path explicitly shown
  - "Cardio" vs "Cardio > Running" clearly distinguished
  - Easier to understand which attributes belong where

PREVIOUS V2.4.8:
- 🎯 PROPER FIX: Excel date column uses DATE format (not TEXT!)
- 🎨 FIXED: Number attributes right-aligned based on data_type

NEW in V2.4.6:
- 🎯 SMART VALIDATION: Auto-reclassify invalid event_ids to CREATE (no manual cleanup!)
- ⚡ PERFORMANCE: Batch query for event validation (1 DB call instead of N)
- 📊 SUMMARY WARNINGS: User-friendly messages (not per-record errors)
- 💾 IMPORT ARCHIVE: Generate _imported.xlsx with corrected event_ids
- 🛡️ DATA SAFETY: Prevent corruption when event_id belongs to wrong category

CRITICAL FIX in V2.4.5:
- 🎯 PROPER FIX: Legend = Source of Truth approach
- ✅ PRINCIPLE: User CAN delete columns, MUST update Legend accordingly
- ❌ REJECT: Import rejected if headers don't match Legend (not auto-ignored!)
- 📝 INSTRUCTIONS: Clear error message explains how to fix Legend
- 💡 FLEXIBLE: User can delete many columns, just needs to update Legend
- 🛡️ SAFE: No silent data corruption, user maintains control

REVERTED from V2.4.4:
- ❌ V2.4.4 auto-ignored mismatches (wrong approach - loses data silently)
- ✅ V2.4.5 rejects with instructions (correct approach - user has control)

CRITICAL FIX in V2.4.3:
- 🐛 FIXED: Off-by-one error in EVENT DATA parsing
- ✅ ISSUE: Parser was looking for headers at row+2 instead of row+1
- ✅ RESULT: Import now correctly reads events from exported Excel files
- 🎯 ROOT CAUSE: Comment said "skip SUBTOTAL row" but export doesn't create one
- 📝 IMPACT: "No events found" error when importing valid Excel files

CRITICAL FIX in V2.4.2:
- 🎯 SOLUTION: First row of each group = SEPARATOR (outside group)
- ✅ RESULT: Multiple DISTINCT collapsible groups in Excel!
- 💡 HOW: Separator row stays visible, breaks group continuity
- 🎨 VISUAL: Separator rows have BOLD text + darker pink background
- 🚀 BENEFIT: Can collapse groups individually, see EVENT DATA header!

PREVIOUS ATTEMPTS:
- V2.4.1: Used group() but Excel merged them (no separators)
- V2.4.0: Used outline_level=1 on all rows (created one big group)

NEW in V2.4:
- ✅ NEW: Legend validation - check if columns from legend exist in Excel
- ✅ NEW: Auto-detect deleted columns with warnings
- ✅ NEW: Report orphan columns (in EVENT DATA but not in legend)
- 🛡️ IMPROVED: Safer import with validation feedback

NEW in V2.3:
- ✅ IMPROVED: Smart row grouping (5-20 rows per group based on total attributes)
- ✅ FIXED: Merged cell formatting - consistent row height, text alignment top
- ✅ FIXED: No more "escaped" text or random row height increases

NEW in V2.2:
- ✅ FIXED: Default/Min/Max values now properly formatted as numbers
- ✅ FIXED: Row grouping in ATTRIBUTE LEGEND (collapsed groups with +/- icons)
- ✅ FIXED: Column grouping F-I (Default/Min/Max/Unit collapsible)
- ✅ FIXED: Comment text vertical alignment with auto row height
- ✅ FIXED: Border on all merged cells (E-I)
- 🎯 NEW: Legend-based import - users can delete legend rows/columns flexibly
- 🎯 NEW: Import maps by column letter from legend, NOT by header names

Excel Format V2.2:
┌─────────────────────────────────────────────────────────────────┐
│ A1: ATTRIBUTE LEGEND (bold title)                               │
├─────┬──────┬──────────────┬───────────┬──────┬─────┬─────┬─────┤
│ Col │ Area │ Category_Path│ Attribute │ Type │ Def │ Min │ Max │ ← F-I grouped
├─────┴──────┴──────────────┴───────────┴──────┴─────┴─────┴─────┤
│ (rows grouped in chunks of 5-7, collapsed, +/- icons ABOVE)     │
├─────────────────────────────────────────────────────────────────┤
│ (empty row)                                                     │
├─────────────────────────────────────────────────────────────────┤
│ EVENT DATA:    Summ (if relevant) ->    [SUBTOTAL formulas]     │
├──────────┬──────┬──────────────┬────────────┬───────────────────┤
│ event_id │ Area │ Category_Path│ event_date │ comment (merged)  │
├──────────┼──────┼──────────────┼────────────┼───────────────────┤
│ 🟣uuid   │ 🟣   │ 🟣           │ 🔵date     │ 🔵 (E:I merged)   │
└──────────┴──────┴──────────────┴────────────┴───────────────────┘

Colors:
🟣 PINK = Read-only (event_id, Area, Category_Path)
🔵 BLUE = Editable (event_date, comment, relevant attributes)
🟠 ORANGE = Non-relevant attribute for this event's category

LEGEND-BASED IMPORT (V2.4.5):
- ATTRIBUTE LEGEND = SOURCE OF TRUTH for column mapping
- Users CAN delete rows from ATTRIBUTE LEGEND (removes attribute from import)
- Users CAN delete columns from EVENT DATA (Excel shifts remaining columns)
- Users MUST update Legend after deleting columns to match new positions
- Import maps columns by LETTER from legend (Col F, Col G...), NOT by header names
- If Legend doesn't match headers, import is REJECTED with instructions
- This allows maximum flexibility while maintaining data integrity

Dependencies: openpyxl, pandas, json
"""

import io
import json
from datetime import datetime, date
from typing import Dict, List, Optional, Tuple, Any, Set
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.datavalidation import DataValidation


# ============================================
# CONSTANTS
# ============================================

# Colors
PINK_FILL = PatternFill(start_color="FFE6F0", end_color="FFE6F0", fill_type="solid")
BLUE_FILL = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # Non-relevant attributes
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
LEGEND_HEADER_FILL = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
TITLE_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=12)
NORMAL_FONT = Font()

BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

OUTER_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Fixed columns for EVENT DATA (before attributes)
FIXED_COLUMNS = ['event_id', 'Area', 'Category_Path', 'event_date', 'session_start', 'comment']
FIXED_COL_COUNT = len(FIXED_COLUMNS)
PADDING_COLS = 3  # Empty columns after comment for grouping padding

# Legend columns
LEGEND_COLUMNS = ['Col', 'Area', 'Category_Path', 'Attribute', 'Type', 'Default', 'Min', 'Max', 'Unit']


# ============================================
# DATA LOADING HELPERS
# ============================================

def load_areas_dict(client, user_id: str) -> Dict[str, Dict]:
    """Load areas as dict: area_id -> {name, ...}"""
    try:
        resp = client.table('areas') \
            .select('id, name, sort_order') \
            .eq('user_id', user_id) \
            .order('sort_order') \
            .execute()
        return {a['id']: a for a in (resp.data or [])}
    except Exception:
        return {}


def load_categories_dict(client, user_id: str) -> Dict[str, Dict]:
    """Load categories as dict: category_id -> {name, full_path, area_id, area_name, ...}"""
    try:
        resp = client.table('categories') \
            .select('id, name, parent_category_id, area_id, level, sort_order') \
            .eq('user_id', user_id) \
            .order('level') \
            .order('sort_order') \
            .execute()
        
        categories = resp.data or []
        cat_dict = {c['id']: c for c in categories}
        areas = load_areas_dict(client, user_id)
        
        result = {}
        for cat in categories:
            path_parts = []
            current = cat
            while current:
                path_parts.insert(0, current['name'])
                parent_id = current.get('parent_category_id')
                current = cat_dict.get(parent_id) if parent_id else None
            
            area_id = cat.get('area_id')
            area_info = areas.get(area_id, {})
            
            # V2.5.2: Include parent_category_id and sort_order for hierarchy operations
            result[cat['id']] = {
                'id': cat['id'],
                'name': cat['name'],
                'full_path': ' > '.join(path_parts),
                'area_id': area_id,
                'area_name': area_info.get('name', 'Unknown'),
                'level': cat.get('level', 1),
                'parent_category_id': cat.get('parent_category_id'),  # V2.5.2: CRITICAL FIX
                'sort_order': cat.get('sort_order', 0)  # V2.5.2: For hierarchical sorting
            }
        
        return result
    except Exception:
        return {}


def get_category_ids_for_area(client, user_id: str, area_id: str) -> List[str]:
    """Get all category IDs belonging to an area."""
    try:
        resp = client.table('categories') \
            .select('id') \
            .eq('user_id', user_id) \
            .eq('area_id', area_id) \
            .execute()
        return [c['id'] for c in (resp.data or [])]
    except Exception:
        return []


def get_hierarchy_levels_for_path(category_path: str, categories_dict: Dict[str, Dict]) -> List[Tuple[str, str]]:
    """
    Extract all hierarchy levels from a Category_Path.
    
    V2.5.3 NEW: Support multi-level event creation from single Excel row.
    
    Example:
        category_path = "Cardio > Running"
        Returns: [
            ("Cardio", category_id_for_cardio),
            ("Cardio > Running", category_id_for_running)
        ]
    
    Args:
        category_path: Full category path (e.g., "Cardio > Running")
        categories_dict: Dict mapping category_id → category info
    
    Returns:
        List of (partial_path, category_id) tuples for each hierarchy level
    """
    if not category_path:
        return []
    
    # Build reverse mapping: full_path → category_id
    path_to_id = {info['full_path']: cat_id for cat_id, info in categories_dict.items()}
    
    # Split path into parts
    parts = [p.strip() for p in category_path.split('>')]
    
    # Build all partial paths
    result = []
    for i in range(1, len(parts) + 1):
        partial_path = ' > '.join(parts[:i])
        category_id = path_to_id.get(partial_path)
        if category_id:
            result.append((partial_path, category_id))
    
    return result


def get_all_descendant_category_ids(categories_dict: Dict[str, Dict], parent_category_ids: List[str]) -> List[str]:
    """
    Get all descendant category IDs for given parent categories.
    
    V2.5.1: When user selects "Cardio", include all children (Running, Cycling, Swimming)
    but NOT siblings (Strength, etc.).
    
    Args:
        categories_dict: Dict mapping category_id to category info
        parent_category_ids: List of parent category IDs to start from
    
    Returns:
        List of all category IDs (parents + all descendants)
    """
    result = set(parent_category_ids)
    
    # Find all descendants recursively
    to_process = list(parent_category_ids)
    while to_process:
        current_id = to_process.pop(0)
        # Find all children of current category
        for cat_id, cat_info in categories_dict.items():
            if cat_info.get('parent_category_id') == current_id:
                if cat_id not in result:
                    result.add(cat_id)
                    to_process.append(cat_id)  # Process children of this child
    
    return list(result)


def get_category_ids_with_parents(categories_dict: Dict[str, Dict], category_ids: List[str]) -> List[str]:
    """
    Get category IDs including all parent categories in their paths.
    
    V2.5.0: This ensures we load attributes from parent categories too.
    Example: For "Cardio > Running", also include "Cardio" attributes.
    
    Args:
        categories_dict: Dict mapping category_id to category info (with parent_category_id)
        category_ids: List of current category IDs
    
    Returns:
        List of category IDs including all parents
    """
    all_ids = set(category_ids)
    
    for cat_id in category_ids:
        cat = categories_dict.get(cat_id)
        # Walk up the parent chain
        while cat and cat.get('parent_category_id'):
            parent_id = cat['parent_category_id']
            all_ids.add(parent_id)
            cat = categories_dict.get(parent_id)
    
    return list(all_ids)


def load_attribute_definitions_for_categories(
    client, user_id: str, category_ids: List[str]
) -> List[Dict]:
    """
    Load all attribute definitions for given categories, sorted hierarchically.
    
    V2.5.2: TRUE hierarchical sort - maintains parent-child tree structure
    Matches Interactive Structure Viewer order exactly!
    """
    if not category_ids:
        return []
    
    try:
        # STEP 1: Load categories with level, sort_order, and parent info
        cats_resp = client.table('categories') \
            .select('id, name, parent_category_id, level, sort_order') \
            .eq('user_id', user_id) \
            .in_('id', category_ids) \
            .execute()
        
        cats_dict = {c['id']: c for c in (cats_resp.data or [])}
        
        # STEP 1.5: Build hierarchical sort paths for each category
        # V2.5.2: This ensures tree structure is maintained
        # Example: "Strength" (sort=1) → "0001"
        #          "Strength > Legs" (sort=1) → "0001/0001"
        #          "Strength > Arms" (sort=2) → "0001/0002"
        #          "Cardio" (sort=2) → "0002"
        
        def get_sort_path(cat_id):
            """Build hierarchical sort path by walking up parent chain."""
            path_parts = []
            current_id = cat_id
            
            while current_id:
                cat = cats_dict.get(current_id)
                if not cat:
                    break
                # Pad sort_order to 4 digits for proper string sorting
                path_parts.insert(0, f"{cat.get('sort_order', 0):04d}")
                current_id = cat.get('parent_category_id')
            
            # Join with '/' to create hierarchical path
            return '/'.join(path_parts)
        
        # STEP 2: Load attribute definitions
        attrs_resp = client.table('attribute_definitions') \
            .select('id, category_id, name, data_type, unit, is_required, default_value, validation_rules, sort_order') \
            .eq('user_id', user_id) \
            .in_('category_id', category_ids) \
            .execute()
        
        attrs = attrs_resp.data or []
        
        # STEP 3: Sort by hierarchical path, then attribute sort_order
        # V2.5.2: This maintains tree structure exactly like Structure Viewer
        def sort_key(attr):
            cat_id = attr.get('category_id')
            sort_path = get_sort_path(cat_id)
            attr_sort = attr.get('sort_order', 999)
            
            # Sort by: hierarchical path (string comparison), then attribute sort
            return (sort_path, attr_sort)
        
        attrs.sort(key=sort_key)
        
        return attrs
        
    except Exception:
        return []


def load_events_for_export(
    client, user_id: str,
    category_ids: Optional[List[str]] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    sort_order: str = 'desc'  # V2.5.5: User-selected sort order
) -> List[Dict]:
    """
    Load events with their attributes for export.
    
    V2.5.5: Added sort_order parameter to respect UI sorting
    V2.5.0: Includes session_start field
    """
    try:
        select_fields = 'id, category_id, event_date, session_start, comment, event_attributes(id, attribute_definition_id, value_text, value_number, value_datetime, value_boolean)'
        
        query = client.table('events') \
            .select(select_fields) \
            .eq('user_id', user_id)
        
        if category_ids:
            query = query.in_('category_id', category_ids)
        
        if date_from:
            query = query.gte('event_date', date_from.isoformat())
        if date_to:
            query = query.lte('event_date', date_to.isoformat())
        
        # V2.5.5: Apply user-selected sort order
        desc_order = (sort_order == 'desc')
        query = query.order('event_date', desc=desc_order) \
                     .order('session_start', desc=desc_order)
        
        resp = query.execute()
        return resp.data or []
    except Exception:
        return []


def merge_session_events(events: List[Dict], categories_dict: Dict[str, Dict]) -> List[Dict]:
    """
    Merge hierarchical events with same session_start AND comment into single export rows.
    
    V2.5.6 FIXED: Group by (timestamp, comment) combination, not just timestamp.
    
    Logic:
    - Group events by (exact session_start timestamp, comment)
    - For each group with multiple events:
      - Check if they form a parent-child hierarchy (different levels)
      - If yes → export only the DEEPEST (leaf) event
      - Merge all attributes from parent events into leaf event
      - If no → export separately (they are independent events)
    
    Example:
    Input:
        Event 1: Cardio (level=1, ts=09:00, comment="novi 2") → [total_duration=5]
        Event 2: Running (level=2, ts=09:00, comment="novi 2") → [distance=10]
    
    Output:
        Merged Event: Running (level=2, ts=09:00, comment="novi 2") → [total_duration=5, distance=10]
    
    Args:
        events: List of event dicts from database
        categories_dict: Category info including level
    
    Returns:
        List of events with merged attributes
    """
    # V2.5.6 FIX: Group by (timestamp, comment) combination
    # This prevents merging unrelated events with same timestamp but different comments
    sessions = {}
    for event in events:
        # Group key: (session_start, comment)
        key = (event.get('session_start'), event.get('comment'))
        if key not in sessions:
            sessions[key] = []
        sessions[key].append(event)
    
    merged_events = []
    
    for (timestamp, comment), session_events in sessions.items():
        if len(session_events) == 1:
            # Single event - no merging needed
            merged_events.append(session_events[0])
            continue
        
        # Multiple events with same timestamp AND comment
        # Get category levels for each event
        event_levels = []
        for event in session_events:
            cat_id = event.get('category_id')
            cat_info = categories_dict.get(cat_id, {})
            level = cat_info.get('level', 0)
            event_levels.append((event, level))
        
        # Sort by level (parent first)
        event_levels.sort(key=lambda x: x[1])
        
        # Check if levels are all different (hierarchical chain)
        levels = [el[1] for el in event_levels]
        unique_levels = set(levels)
        
        if len(unique_levels) == len(levels) and len(unique_levels) > 1:
            # All different levels - this is a parent-child hierarchy
            # Merge into the deepest (leaf) event
            
            leaf_event = event_levels[-1][0].copy()  # Deepest level
            
            # Merge all attributes from all events in the session
            merged_attributes = []
            for event, _ in event_levels:
                merged_attributes.extend(event.get('event_attributes', []))
            
            # Replace leaf event's attributes with merged list
            leaf_event['event_attributes'] = merged_attributes
            
            merged_events.append(leaf_event)
        else:
            # Not all different levels - these are independent events
            # (e.g., 2 Cardio events at same time, or same-level siblings)
            # Export separately
            merged_events.extend(session_events)
    
    return merged_events


def parse_validation_rules(rules) -> Dict:
    """Safely parse validation_rules which can be dict, JSON string, or None."""
    if isinstance(rules, dict):
        return rules
    if isinstance(rules, str):
        try:
            return json.loads(rules)
        except:
            return {}
    return {}


# ============================================
# EXCEL EXPORT V2
# ============================================

def create_events_excel_v2(
    events: List[Dict],
    attribute_definitions: List[Dict],
    categories_dict: Dict[str, Dict]
) -> bytes:
    """Create Excel file with enhanced V2.5 format."""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Events"
    
    # Build attribute info and determine columns
    # V2.5.0: attr_columns now contains (category_path, attr_name, attr_def_id) tuples for uniqueness
    attr_info = {}  # attr_def_id -> {name, category_path, area_name, data_type, ...}
    attr_columns = []  # Ordered list of (category_path, attr_name, attr_def_id) tuples
    attr_by_category = {}  # category_id -> set of attr_def_ids
    
    for attr_def in attribute_definitions:
        cat_id = attr_def.get('category_id')
        cat_info = categories_dict.get(cat_id, {})
        validation = parse_validation_rules(attr_def.get('validation_rules'))
        
        category_path = cat_info.get('full_path', 'Unknown')
        attr_name = attr_def['name']
        attr_def_id = attr_def['id']
        
        attr_info[attr_def_id] = {
            'id': attr_def_id,
            'name': attr_name,
            'category_id': cat_id,
            'category_path': category_path,
            'area_name': cat_info.get('area_name', 'Unknown'),
            'data_type': attr_def.get('data_type', 'text'),
            'unit': attr_def.get('unit', ''),
            'default_value': attr_def.get('default_value', ''),
            'min': validation.get('min', ''),
            'max': validation.get('max', '')
        }
        
        # V2.5.0: Use (category_path, attr_name, attr_def_id) as unique key
        # This prevents losing attributes with duplicate names from different categories
        attr_unique_key = (category_path, attr_name, attr_def_id)
        if attr_unique_key not in attr_columns:
            attr_columns.append(attr_unique_key)
        
        if cat_id not in attr_by_category:
            attr_by_category[cat_id] = set()
        attr_by_category[cat_id].add(attr_def_id)
    
    # Build category -> attr_names mapping for orange highlighting
    cat_to_attr_names = {}
    for cat_id, attr_ids in attr_by_category.items():
        cat_to_attr_names[cat_id] = {attr_info[aid]['name'] for aid in attr_ids}
    
    # ─────────────────────────────────────────
    # SECTION 1: ATTRIBUTE LEGEND
    # ─────────────────────────────────────────
    
    # Set outline settings: summary rows/columns ABOVE/LEFT (not below/right)
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = False
    
    row = 1
    
    # Title row
    ws.cell(row=row, column=1, value="ATTRIBUTE LEGEND:").font = TITLE_FONT
    row += 1
    
    # Header row
    legend_header_row = row
    for col_idx, col_name in enumerate(LEGEND_COLUMNS, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.fill = LEGEND_HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    row += 1
    
    # Legend data rows (grouped in smart chunks based on total count)
    legend_start_row = row
    attr_col_start = FIXED_COL_COUNT + PADDING_COLS + 1
    
    # First pass: write all legend data
    # V2.5.0: attr_columns now contains (category_path, attr_name, attr_def_id) tuples
    legend_rows = []
    for idx, (category_path, attr_name, attr_def_id) in enumerate(attr_columns):
        info = attr_info.get(attr_def_id, {})
        col_letter = get_column_letter(attr_col_start + idx)
        
        # Convert default, min, max to numbers if data_type is number
        data_type = info.get('data_type', 'text')
        default_val = info.get('default_value', '')
        min_val = info.get('min', '')
        max_val = info.get('max', '')
        
        # Convert to numbers for number type
        if data_type == 'number':
            if default_val:
                try:
                    default_val = float(default_val) if '.' in str(default_val) else int(default_val)
                except (ValueError, TypeError):
                    pass
            if min_val:
                try:
                    min_val = float(min_val) if '.' in str(min_val) else int(min_val)
                except (ValueError, TypeError):
                    pass
            if max_val:
                try:
                    max_val = float(max_val) if '.' in str(max_val) else int(max_val)
                except (ValueError, TypeError):
                    pass
        
        legend_data = [
            col_letter,
            info.get('area_name', ''),
            info.get('category_path', ''),
            attr_name,
            data_type,
            default_val,
            min_val,
            max_val,
            info.get('unit', '')
        ]
        
        for col_idx, value in enumerate(legend_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value if value != '' else '')
            cell.fill = PINK_FILL
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")
        
        legend_rows.append(row)
        row += 1
    
    legend_end_row = row - 1
    
    # Second pass: create smart groups with SEPARATORS
    # CRITICAL: First row of each group stays OUTSIDE the group as separator
    # This creates distinct collapsible groups in Excel!
    if legend_rows:
        total_attrs = len(legend_rows)
        
        if total_attrs <= 5:
            # Single group for small datasets - no separators needed
            num_groups = 1
        elif total_attrs <= 100:
            # Target ~10 rows per group for medium datasets
            num_groups = max(1, (total_attrs + 9) // 10)
        else:
            # Large datasets - max 20 rows per group
            num_groups = max(5, (total_attrs + 19) // 20)
        
        # Calculate actual group size (distribute evenly)
        actual_group_size = (total_attrs + num_groups - 1) // num_groups
        
        # FIXED: Use FIRST row as separator (outside group), group the REST
        # This creates SEPARATE groups because separators break the continuity
        for g in range(num_groups):
            start_idx = g * actual_group_size
            end_idx = min(start_idx + actual_group_size - 1, total_attrs - 1)
            
            if start_idx <= end_idx and end_idx < len(legend_rows):
                # First row of this group = SEPARATOR (stays outside group)
                separator_row = legend_rows[start_idx]
                
                # Make separator BOLD and slightly different background for visibility
                for col_idx in range(1, 10):  # Columns A-I
                    cell = ws.cell(row=separator_row, column=col_idx)
                    cell.font = Font(bold=True)
                    # Slightly darker pink for separator
                    cell.fill = PatternFill(start_color="FFD0E0", end_color="FFD0E0", fill_type="solid")
                
                # Group ONLY the rows AFTER separator (if there are any)
                if end_idx > start_idx:  # More than 1 row in this group
                    group_start = legend_rows[start_idx + 1]  # Skip separator
                    group_end = legend_rows[end_idx]
                    
                    if group_end > group_start:  # Only group if more than 1 row
                        try:
                            ws.row_dimensions.group(group_start, group_end, hidden=True)
                        except Exception as e:
                            # Fallback: set outline level
                            for row_num in range(group_start, group_end + 1):
                                ws.row_dimensions[row_num].outline_level = 1
                                ws.row_dimensions[row_num].hidden = True
    
    # Column grouping for Default, Min, Max, Unit (columns F-I)
    # Set outline_level explicitly for each column
    for col_letter in ['F', 'G', 'H', 'I']:
        ws.column_dimensions[col_letter].outline_level = 1
    
    # ─────────────────────────────────────────
    # EMPTY ROW
    # ─────────────────────────────────────────
    row += 1
    
    # ─────────────────────────────────────────
    # SECTION 2: EVENT DATA
    # ─────────────────────────────────────────
    
    # Title row with SUBTOTAL placeholders
    event_title_row = row
    ws.cell(row=row, column=1, value="EVENT DATA:").font = TITLE_FONT
    ws.cell(row=row, column=3, value="Summ (if relevant) ->").alignment = Alignment(horizontal="right")
    
    # SUBTOTAL formulas will be added after we know the data range
    row += 1
    
    # Header row
    # V2.5.0: Build headers with "attr_name (Category)" format
    event_header_row = row
    
    # Create attribute header strings with category name for clarity
    attr_header_strings = []
    for category_path, attr_name, attr_def_id in attr_columns:
        # Extract last part of category_path as short name
        category_short = category_path.split(' > ')[-1] if ' > ' in category_path else category_path
        header_str = f"{attr_name} ({category_short})"
        attr_header_strings.append(header_str)
    
    all_columns = FIXED_COLUMNS + [''] * PADDING_COLS + attr_header_strings
    
    for col_idx, col_name in enumerate(all_columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    
    row += 1
    event_data_start_row = row
    
    # ─────────────────────────────────────────
    # EVENT DATA ROWS
    # ─────────────────────────────────────────
    
    for event in events:
        event_id = event.get('id', '')
        cat_id = event.get('category_id')
        cat_info = categories_dict.get(cat_id, {})
        
        # V2.5.0: Get all relevant attribute IDs for this event's category
        # Include attributes from this category AND all parent categories
        relevant_attr_ids = set()
        current_cat_id = cat_id
        while current_cat_id:
            if current_cat_id in attr_by_category:
                relevant_attr_ids.update(attr_by_category[current_cat_id])
            # Walk up to parent
            cat = categories_dict.get(current_cat_id, {})
            current_cat_id = cat.get('parent_category_id')
        
        # Build attribute values dict: (category_path, attr_name, attr_def_id) -> value
        attr_values = {}
        for ea in event.get('event_attributes', []):
            attr_def_id = ea.get('attribute_definition_id')
            attr_inf = attr_info.get(attr_def_id, {})
            
            if attr_inf:
                # Build key matching attr_columns format
                key = (attr_inf['category_path'], attr_inf['name'], attr_def_id)
                
                value = None
                if ea.get('value_number') is not None:
                    value = ea['value_number']
                elif ea.get('value_boolean') is not None:
                    value = ea['value_boolean']
                elif ea.get('value_datetime'):
                    value = ea['value_datetime']
                elif ea.get('value_text'):
                    value = ea['value_text']
                attr_values[key] = value
        
        # Fixed columns
        # V2.4.8: Convert event_date to Python date object for Excel
        event_date = event.get('event_date', '')
        if isinstance(event_date, str) and event_date:
            try:
                event_date = datetime.strptime(event_date, '%Y-%m-%d').date()
            except:
                pass  # Keep as string if parsing fails
        elif isinstance(event_date, datetime):
            event_date = event_date.date()
        # If already date object, keep it
        
        # V2.5.0: Handle session_start (TIME column)
        session_start = event.get('session_start', '')
        if session_start:
            # Parse timestamp to time
            if isinstance(session_start, str):
                try:
                    dt = datetime.fromisoformat(session_start.replace('Z', '+00:00'))
                    session_start = dt.strftime('%H:%M')
                except:
                    session_start = '09:00'  # Default
            elif isinstance(session_start, datetime):
                session_start = session_start.strftime('%H:%M')
        else:
            session_start = '09:00'  # Default if NULL
        
        fixed_data = [
            event_id,
            cat_info.get('area_name', ''),
            cat_info.get('full_path', ''),
            event_date,
            session_start,  # V2.5.0: TIME column
            event.get('comment', '') or ''
        ]
        
        # Write fixed columns (A-E: event_id, Area, Category_Path, event_date, session_start)
        for col_idx, value in enumerate(fixed_data[:5], start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Color: event_id(1), Area(2), Category_Path(3) = PINK; event_date(4), session_start(5) = BLUE
            if col_idx <= 3:
                cell.fill = PINK_FILL
            else:
                cell.fill = BLUE_FILL
            
            # Set formats
            if col_idx == 4:  # event_date column
                cell.number_format = 'YYYY-MM-DD'  # Excel DATE format with ISO display
            elif col_idx == 5:  # session_start (TIME) column
                cell.number_format = '@'  # Text format for HH:MM
        
        # Comment column (F) with merge to padding columns (F:H after we reduced PADDING_COLS to 3)
        # V2.5.0: Column indices shifted due to adding session_start
        comment_value = fixed_data[5]
        comment_start_col = 6  # Column F
        comment_end_col = comment_start_col + PADDING_COLS  # Merge to col I (6+3-1=8, so F:I)
        ws.merge_cells(f'{get_column_letter(comment_start_col)}{row}:{get_column_letter(comment_end_col)}{row}')
        comment_cell = ws.cell(row=row, column=comment_start_col, value=comment_value)
        comment_cell.fill = BLUE_FILL
        comment_cell.border = BORDER
        
        # IMPORTANT: vertical="top" prevents text from "escaping" too high
        # NO wrap_text by default - keeps rows consistent height
        comment_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)
        
        # Apply border to ALL cells in merged range
        for col_idx in range(comment_start_col, comment_end_col + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = BORDER
            cell.fill = BLUE_FILL
        
        # Set DEFAULT row height for all event data rows (consistent look)
        # This prevents automatic height adjustment which causes inconsistent appearance
        ws.row_dimensions[row].height = 20  # Standard row height
        
        # Only increase height for VERY long comments (>100 chars)
        # and enable wrap_text only in those cases
        if comment_value and len(str(comment_value)) > 100:
            comment_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            # Calculate height based on text length (more conservative than before)
            estimated_lines = min(3, len(str(comment_value)) // 50 + 1)  # Max 3 lines
            ws.row_dimensions[row].height = 20 + (estimated_lines - 1) * 15
        
        # Attribute columns
        # V2.5.0: Use (category_path, attr_name, attr_def_id) as key
        for attr_idx, attr_key in enumerate(attr_columns):
            col_idx = attr_col_start + attr_idx
            category_path, attr_name, attr_def_id = attr_key
            value = attr_values.get(attr_key, '')
            
            cell = ws.cell(row=row, column=col_idx, value=value if value is not None else '')
            cell.border = BORDER
            
            # V2.5.1 FIX: Better color logic - check if attr's category is in event's hierarchy
            # BLUE if attr belongs to event's category OR any parent in hierarchy
            # ORANGE if attr belongs to different branch
            attr_cat_id = attr_info.get(attr_def_id, {}).get('category_id')
            is_relevant = False
            
            # Walk up event's category hierarchy and check if attr's category matches
            test_cat_id = cat_id
            while test_cat_id:
                if attr_cat_id == test_cat_id:
                    is_relevant = True
                    break
                cat_test = categories_dict.get(test_cat_id, {})
                test_cat_id = cat_test.get('parent_category_id')
            
            if is_relevant:
                cell.fill = BLUE_FILL
            else:
                cell.fill = ORANGE_FILL
            
            # V2.4.8: Check data_type for alignment
            attr_inf = attr_info.get(attr_def_id)
            if attr_inf and attr_inf.get('data_type') == 'number':
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if isinstance(value, (int, float)):
                    cell.number_format = '0.##'
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        row += 1
    
    event_data_end_row = row - 1
    
    # ─────────────────────────────────────────
    # SUBTOTAL FORMULAS
    # ─────────────────────────────────────────
    
    # V2.5.0: attr_columns now contains (category_path, attr_name, attr_def_id) tuples
    for attr_idx, (category_path, attr_name, attr_def_id) in enumerate(attr_columns):
        # Check if it's a number type
        attr_inf = attr_info.get(attr_def_id)
        if attr_inf and attr_inf.get('data_type') == 'number':
            col_idx = attr_col_start + attr_idx
            col_letter = get_column_letter(col_idx)
            
            # SUBTOTAL(9, range) = SUM that respects filters
            formula = f"=SUBTOTAL(9,{col_letter}{event_data_start_row}:{col_letter}{event_data_end_row})"
            cell = ws.cell(row=event_title_row, column=col_idx, value=formula)
            cell.alignment = Alignment(horizontal="right")
    
    # ─────────────────────────────────────────
    # AUTOFILTER
    # ─────────────────────────────────────────
    
    # AutoFilter on header row, columns A through last column
    last_col = len(all_columns)
    ws.auto_filter.ref = f"A{event_header_row}:{get_column_letter(last_col)}{event_data_end_row}"
    
    # ─────────────────────────────────────────
    # FREEZE PANES
    # ─────────────────────────────────────────
    
    # V2.5.0: Freeze below header row, after session_start column (F = column 6)
    ws.freeze_panes = f"G{event_data_start_row}"
    
    # ─────────────────────────────────────────
    # COLUMN WIDTHS
    # ─────────────────────────────────────────
    
    # V2.5.0: Adjusted for new session_start (TIME) column
    ws.column_dimensions['A'].width = 10  # event_id (narrow)
    ws.column_dimensions['B'].width = 12  # Area
    ws.column_dimensions['C'].width = 30  # Category_Path
    ws.column_dimensions['D'].width = 12  # event_date
    ws.column_dimensions['E'].width = 8   # session_start (TIME)
    ws.column_dimensions['F'].width = 30  # comment (merged F:I)
    
    # Padding columns (G-H are part of comment merge, I is last padding)
    for i in range(1, PADDING_COLS):  # Columns G, H
        ws.column_dimensions[get_column_letter(FIXED_COL_COUNT + i)].width = 3
    
    # Attribute columns
    for idx in range(len(attr_columns)):
        col_letter = get_column_letter(attr_col_start + idx)
        ws.column_dimensions[col_letter].width = 12
    
    # Legend columns widths
    legend_widths = {'A': 6, 'B': 12, 'C': 30, 'D': 15, 'E': 10, 'F': 10, 'G': 8, 'H': 8, 'I': 10}
    for col, width in legend_widths.items():
        current = ws.column_dimensions[col].width or 0
        if current < width:
            ws.column_dimensions[col].width = width
    
    # ─────────────────────────────────────────
    # HELP SHEET
    # ─────────────────────────────────────────
    
    ws_help = wb.create_sheet("Help")
    _create_help_sheet_v2(ws_help)
    
    # Save
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()


def _create_help_sheet_v2(ws):
    """Create Help sheet with V2.5.0 instructions."""
    instructions = [
        ["EVENTS TRACKER - Excel Export/Import Help V2.5.0"],
        [""],
        ["🎯 IMPORTANT: ATTRIBUTE LEGEND = SOURCE OF TRUTH"],
        [""],
        ["The ATTRIBUTE LEGEND tells import which Excel column contains which attribute."],
        ["You MUST keep Legend synchronized with your column structure!"],
        [""],
        ["═══════════════════════════════════════════════════════════════"],
        [""],
        ["🆕 NEW IN V2.5.0:"],
        [""],
        ["✨ TIME Column (session_start) added after event_date"],
        ["   - Default: 09:00 if not specified"],
        ["   - Format: HH:MM (24-hour, e.g., 14:30)"],
        ["   - Allows tracking multiple events per day"],
        [""],
        ["🌳 Parent Category Attributes Included"],
        ["   - Running events now show Cardio parent attributes"],
        ["   - All parent attributes are EDITABLE (blue)"],
        ["   - Example: Running event shows distance, duration (Running) AND"],
        ["     total_duration, avg_hr, max_hr, calories (Cardio parent)"],
        [""],
        ["🏷️ Clearer Attribute Names in Headers"],
        ["   - Headers show 'attr_name (Category)' format"],
        ["   - Example: 'distance (Running)', 'distance (Cycling)'"],
        ["   - Prevents confusion with duplicate attribute names"],
        [""],
        ["📊 Hierarchical Sorting"],
        ["   - Attributes sorted by category hierarchy"],
        ["   - Parent categories appear before child categories"],
        ["   - Matches Interactive Structure Viewer order"],
        [""],
        ["═══════════════════════════════════════════════════════════════"],
        [""],
        ["📋 FILE STRUCTURE:"],
        [""],
        ["1. ATTRIBUTE LEGEND (top section)"],
        ["   - Col: Column letter (J, K, L...) for this attribute in EVENT DATA"],
        ["   - Area: Which area this attribute belongs to"],
        ["   - Category_Path: Full category path (e.g., 'Cardio > Running')"],
        ["   - Attribute: Attribute name"],
        ["   - Type/Default/Min/Max/Unit: Attribute properties"],
        ["   - Rows grouped (click +/- ABOVE group to expand/collapse)"],
        [""],
        ["2. EVENT DATA (bottom section)"],
        ["   - Your actual events with attribute values"],
        ["   - Fixed columns: event_id, Area, Category_Path, event_date, session_start"],
        ["   - Comment column merged (F:I) for more space"],
        ["   - Attribute columns start at J with 'name (Category)' headers"],
        ["   - AutoFilter enabled - click headers to filter"],
        ["   - Title row shows SUMs (respects filters)"],
        [""],
        ["═══════════════════════════════════════════════════════════════"],
        [""],
        ["🎨 COLOR CODING:"],
        [""],
        ["🟣 PINK = READ-ONLY (do not edit)"],
        ["   - event_id: UUID identifying existing events"],
        ["   - Area: Auto-determined by Category"],
        ["   - Category_Path: Cannot change existing event's category"],
        [""],
        ["🔵 BLUE = EDITABLE"],
        ["   - event_date: Date (YYYY-MM-DD format)"],
        ["   - session_start: Time (HH:MM, default 09:00)"],
        ["   - comment: Notes (merged F:I for space)"],
        ["   - Attributes relevant for this category AND parent categories"],
        [""],
        ["🟠 ORANGE = NOT RELEVANT"],
        ["   - Attribute belongs to different category (not in hierarchy)"],
        ["   - Can leave empty - will be ignored"],
        [""],
        ["═══════════════════════════════════════════════════════════════"],
        [""],
        ["✏️ HOW TO EDIT EVENTS:"],
        [""],
        ["UPDATE EXISTING EVENTS:"],
        ["1. Find row with event_id filled (UUID in column A)"],
        ["2. Change BLUE columns only (date, time, comment, attributes)"],
        ["3. Save and import"],
        [""],
        ["CREATE NEW EVENTS:"],
        ["1. Add row at bottom, leave event_id EMPTY"],
        ["2. Fill Area, Category_Path (must exist in your structure)"],
        ["3. Fill event_date (required, YYYY-MM-DD format)"],
        ["4. Fill session_start (optional, HH:MM format, defaults to 09:00)"],
        ["5. Fill attribute values (only relevant ones - blue cells)"],
        ["6. Save and import"],
        [""],
        ["═══════════════════════════════════════════════════════════════"],
        [""],
        ["✂️ HOW TO REMOVE ATTRIBUTES - TWO OPTIONS:"],
        [""],
        ["OPTION 1: Delete Legend Rows (SIMPLEST)"],
        ["  1. Open ATTRIBUTE LEGEND section"],
        ["  2. Find attribute you don't want (e.g., 'Weight' in row 5)"],
        ["  3. DELETE entire row from ATTRIBUTE LEGEND"],
        ["  4. DON'T touch EVENT DATA columns"],
        ["  5. Save and import → Attribute ignored ✅"],
        [""],
        ["OPTION 2: Delete Columns and Update Legend"],
        ["  1. DELETE unwanted columns from EVENT DATA (e.g., Col K)"],
        ["  2. Excel automatically shifts remaining columns LEFT"],
        ["  3. UPDATE 'Col' letters in ATTRIBUTE LEGEND to match NEW positions"],
        ["     Example: If L shifted to K, change 'Col L' to 'Col K'"],
        ["  4. OR DELETE legend rows for removed columns"],
        ["  5. Save and import → Works perfectly! ✅"],
        [""],
        ["⚠️ CRITICAL: If you delete columns WITHOUT updating Legend:"],
        ["   → Import will FAIL with error message"],
        ["   → Error will list mismatched columns"],
        ["   → You must fix Legend before importing"],
        [""],
        ["═══════════════════════════════════════════════════════════════"],
        [""],
        ["🔧 FIXING MISMATCH ERRORS:"],
        [""],
        ["If you see: 'Cannot import: Column headers don't match LEGEND!'"],
        [""],
        ["This means Legend Col letters don't match actual column positions."],
        ["This happens when you delete columns but forget to update Legend."],
        [""],
        ["TO FIX:"],
        ["1. Open ATTRIBUTE LEGEND section"],
        ["2. For each mismatched column mentioned in error:"],
        ["   - EITHER: UPDATE 'Col' letter to match new position"],
        ["   - OR: DELETE entire legend row if you don't want it"],
        ["3. Save Excel and import again"],
        [""],
        ["Example:"],
        ["  Original: Col K=duration, Col L=pace, Col M=type"],
        ["  You delete Col K → Excel shifts L to K, M to L"],
        ["  Update Legend: Change 'Col L' to 'Col K' (pace)"],
        ["  Update Legend: Change 'Col M' to 'Col L' (type)"],
        ["  Import succeeds! ✅"],
        [""],
        ["═══════════════════════════════════════════════════════════════"],
        [""],
        ["💡 TIPS:"],
        ["- Use AutoFilter to show only specific categories/dates"],
        ["- Collapse LEGEND groups to see more EVENT DATA"],
        ["- SUM row updates automatically when you filter"],
        ["- Orange cells can be left empty (not relevant)"],
        ["- +/- icons for groups are ABOVE the group"],
        ["- Option 1 (delete legend rows) is SAFEST"],
        ["- Option 2 (delete columns) requires Legend update"],
        [""],
        ["⚠️ IMPORTANT WARNINGS:"],
        ["- DO NOT change event_id values (breaks event tracking)"],
        ["- Empty cells = no value (not zero)"],
        ["- Import maps by LEGEND Col letters, not header names"],
        ["- Delete in app is safer than Excel row deletion"],
        ["- ALWAYS keep Legend synchronized with columns!"],
        [""],
        ["═══════════════════════════════════════════════════════════════"],
        [""],
        ["🎯 KEY PRINCIPLE: ATTRIBUTE LEGEND = SOURCE OF TRUTH"],
        [""],
        ["Import uses Legend to know which column contains which attribute."],
        ["You have full flexibility to organize columns as you want,"],
        ["as long as Legend correctly describes the structure!"],
    ]
    
    for row_idx, row_data in enumerate(instructions, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif value and (value.endswith(':') or value.startswith('🎯')) and not value.startswith(' '):
                cell.font = Font(bold=True, size=11)
    
    ws.column_dimensions['A'].width = 75


# ============================================
# EXCEL IMPORT V2 - LEGEND-BASED MAPPING
# ============================================

def parse_events_excel_v2(file_bytes: bytes) -> Tuple[List[Dict], List[Dict], Dict[str, Tuple[str, str, str]], str]:
    """
    Parse Excel file V2 with LEGEND-BASED column mapping.
    
    This allows users to:
    - Delete rows from ATTRIBUTE LEGEND (removes that attribute from import)
    - Delete columns from EVENT DATA (removes that attribute from import)
    - Import maps columns by letter from legend, NOT by header names
    
    Returns:
        events_to_create: List of new events to create
        events_to_update: List of existing events to update
        legend_mapping: Dict[col_letter -> (area, category_path, attribute_name)]
        error_message: Error string if parsing fails
    """
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        
        # ─────────────────────────────────────────
        # STEP 1: Parse ATTRIBUTE LEGEND
        # ─────────────────────────────────────────
        
        legend_start_row = None
        for row_idx in range(1, min(ws.max_row + 1, 200)):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value and 'ATTRIBUTE LEGEND' in str(cell_value):
                legend_start_row = row_idx
                break
        
        if not legend_start_row:
            return [], [], {}, "Could not find ATTRIBUTE LEGEND section. Invalid file format."
        
        # Legend header row is next row after title
        legend_header_row = legend_start_row + 1
        
        # Build legend mapping: Col letter -> (Area, Category_Path, Attribute)
        legend_mapping = {}
        
        for row_idx in range(legend_header_row + 1, ws.max_row + 1):
            col_cell = ws.cell(row=row_idx, column=1).value  # Col column
            if not col_cell:
                break  # End of legend section
            
            col_letter = str(col_cell).strip().upper()
            area = ws.cell(row=row_idx, column=2).value or ''
            category_path = ws.cell(row=row_idx, column=3).value or ''
            attribute = ws.cell(row=row_idx, column=4).value or ''
            
            # Skip if missing essential info
            if not col_letter or not attribute:
                continue
            
            # Remove "COL" prefix if present (e.g., "COL F" -> "F")
            if col_letter.startswith('COL '):
                col_letter = col_letter[4:].strip()
            
            legend_mapping[col_letter] = (area, category_path, attribute)
        
        if not legend_mapping:
            return [], [], {}, "No valid attribute mappings found in ATTRIBUTE LEGEND."
        
        # ─────────────────────────────────────────
        # STEP 1.5: Validate LEGEND (NEW V2.4)
        # ─────────────────────────────────────────
        # Check if columns from legend exist in the Excel sheet
        
        validation_warnings = []
        max_col_letter = get_column_letter(ws.max_column)
        
        for col_letter in legend_mapping.keys():
            # Check if column exists in sheet (within max_column range)
            try:
                col_idx = column_index_from_string(col_letter)
                if col_idx > ws.max_column:
                    validation_warnings.append(
                        f"⚠️ Column {col_letter} in LEGEND but NOT in EVENT DATA (sheet ends at {max_col_letter}). "
                        f"Attribute '{legend_mapping[col_letter][2]}' will be ignored."
                    )
            except Exception:
                validation_warnings.append(
                    f"⚠️ Invalid column letter '{col_letter}' in LEGEND. Will be ignored."
                )
        
        # Check if there are "orphan" columns in EVENT DATA (not in legend)
        # This is less critical but good to know
        # V2.5.0: Column J starts attributes (after comment merged F-I)
        event_data_start_col = 10  # Column J
        if ws.max_column > event_data_start_col:
            orphan_columns = []
            for col_idx in range(event_data_start_col, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                if col_letter not in legend_mapping:
                    orphan_columns.append(col_letter)
            
            if orphan_columns:
                validation_warnings.append(
                    f"ℹ️ Columns {', '.join(orphan_columns[:5])} in EVENT DATA but NOT in LEGEND. "
                    f"These columns will be ignored during import."
                )
        
        # If there are critical validation issues, return them as error
        # For now, we treat all as warnings and continue with import
        validation_info = "\n".join(validation_warnings) if validation_warnings else ""
        
        # ─────────────────────────────────────────
        # STEP 2: Find EVENT DATA section
        # ─────────────────────────────────────────
        
        event_data_row = None
        for row_idx in range(legend_header_row, min(ws.max_row + 1, 500)):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value and 'EVENT DATA' in str(cell_value):
                event_data_row = row_idx
                break
        
        if not event_data_row:
            return [], [], legend_mapping, "Could not find EVENT DATA section. Invalid file format."
        
        # Header row is ONE row after EVENT DATA title
        # (Export writes headers immediately after title, no subtotal row between)
        header_row = event_data_row + 1
        
        # ─────────────────────────────────────────
        # STEP 2.5: Validate LEGEND vs HEADERS (V2.5.0 updated)
        # ─────────────────────────────────────────
        # V2.5.0: Headers now have "attr_name (Category)" format
        # Check if attribute name appears in header (flexible matching)
        # Legend = SOURCE OF TRUTH: User MUST update legend if columns change
        
        mismatch_errors = []
        
        for col_letter, (area, cat_path, attr_name) in legend_mapping.items():
            try:
                col_idx = column_index_from_string(col_letter)
                if col_idx <= ws.max_column:
                    # Get actual header from EVENT DATA table
                    actual_header = ws.cell(row=header_row, column=col_idx).value
                    actual_header_str = str(actual_header).strip() if actual_header else ''
                    
                    # V2.5.0: Headers have "attr_name (Category)" format
                    # Check if attr_name appears at start of header
                    # This allows both old "attr_name" and new "attr_name (Category)" formats
                    if actual_header_str:
                        # Extract base name before any parentheses
                        base_header = actual_header_str.split('(')[0].strip()
                        if base_header != attr_name:
                            mismatch_errors.append(
                                f"Col {col_letter}: Legend says '{attr_name}' but header shows '{actual_header_str}'"
                            )
            except Exception:
                # Column letter invalid - already handled in previous validation
                pass
        
        # If ANY mismatches found, REJECT import with instructions
        if mismatch_errors:
            error_msg = (
                "❌ Cannot import: Column headers don't match ATTRIBUTE LEGEND!\n\n"
                "This usually happens when you delete columns from EVENT DATA.\n"
                "Excel shifts remaining columns, but Legend still shows old positions.\n\n"
                "🔍 Mismatches found:\n" + 
                "\n".join(f"  • {err}" for err in mismatch_errors) +
                "\n\n"
                "📝 How to fix:\n"
                "1. Open the ATTRIBUTE LEGEND section in Excel\n"
                "2. For each mismatched column, you have TWO options:\n"
                "   a) UPDATE the 'Col' letter to match the new column position, OR\n"
                "   b) DELETE the entire legend row if you don't want to import that attribute\n"
                "3. Save Excel and try import again\n\n"
                "💡 Example: If you deleted Col K and L shifted left to K:\n"
                "   Change 'Col L' to 'Col K' in the legend row\n\n"
                "✅ Remember: ATTRIBUTE LEGEND is the source of truth for import mapping!"
            )
            return [], [], legend_mapping, error_msg
        
        
        # ─────────────────────────────────────────
        # STEP 3: Parse data rows using legend mapping
        # ─────────────────────────────────────────
        
        events_to_create = []
        events_to_update = []
        
        # Build reverse mapping: column_index -> attribute_name (from legend)
        col_to_attr = {}
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            if col_letter in legend_mapping:
                _, _, attr_name = legend_mapping[col_letter]
                col_to_attr[col_idx] = attr_name
        
        for row_idx in range(header_row + 1, ws.max_row + 1):
            # Check for Area in column 2 to detect valid data rows
            area_cell = ws.cell(row=row_idx, column=2).value
            if not area_cell:
                continue
            
            # Build row data
            row_data = {}
            
            # Fixed columns (A-F) - V2.5.0: Added session_start
            row_data['event_id'] = ws.cell(row=row_idx, column=1).value
            row_data['Area'] = ws.cell(row=row_idx, column=2).value
            row_data['Category_Path'] = ws.cell(row=row_idx, column=3).value
            row_data['event_date'] = ws.cell(row=row_idx, column=4).value
            row_data['session_start'] = ws.cell(row=row_idx, column=5).value  # V2.5.0: TIME column
            
            # Comment is merged F-I, read from F (V2.5.0: column shifted due to session_start)
            row_data['comment'] = ws.cell(row=row_idx, column=6).value
            
            # Attribute columns (mapped by legend)
            for col_idx, attr_name in col_to_attr.items():
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is not None:  # Include even empty strings, skip only None
                    row_data[attr_name] = value
            
            # Determine CREATE vs UPDATE based on event_id
            event_id = row_data.get('event_id')
            
            if event_id is None or str(event_id).strip() == '':
                events_to_create.append(row_data)
            else:
                events_to_update.append(row_data)
        
        return events_to_create, events_to_update, legend_mapping, validation_info
        
    except Exception as e:
        return [], [], {}, f"Error parsing Excel file: {str(e)}"


def parse_events_excel(file_bytes: bytes) -> Tuple[List[Dict], List[Dict], str]:
    """
    DEPRECATED: Old import function. Use parse_events_excel_v2() instead.
    
    This is kept for backward compatibility but should not be used.
    The new V2 function uses legend-based mapping which is more flexible.
    """
    # Call V2 and discard legend_mapping for backward compatibility
    creates, updates, _, error = parse_events_excel_v2(file_bytes)
    return creates, updates, error


def smart_reclassify_events(
    client,
    user_id: str,
    events_to_create: List[Dict],
    events_to_update: List[Dict],
    categories_dict: Dict[str, Dict]
) -> Tuple[List[Dict], List[Dict], List[str]]:
    """
    Smart validation of event_ids with auto-reclassification.
    
    V2.4.6 NEW: When users drag Excel rows down, event_id gets copied.
    This causes problems:
    1. event_id doesn't exist → UPDATE fails silently
    2. event_id exists BUT wrong category → DATA CORRUPTION!
    
    Solution:
    - Batch query all event_ids (performance - 1 DB call!)
    - Check: exists + category match
    - Auto-convert invalid → CREATE (safer than failing!)
    - Show summary warning (not per-record errors)
    
    User doesn't need to manually clear event_ids! 🎉
    
    Args:
        client: Supabase client
        user_id: Current user ID
        events_to_create: List of events without event_id
        events_to_update: List of events with event_id
        categories_dict: Dict mapping category_id → info
        
    Returns:
        Tuple of:
        - Updated events_to_create (with reclassified events)
        - Updated events_to_update (only valid updates)
        - List of warning messages
    """
    if not events_to_update:
        return events_to_create, events_to_update, []
    
    warnings = []
    
    # Build reverse mapping: Category_Path → category_id
    cat_by_path = {info['full_path']: cat_id for cat_id, info in categories_dict.items()}
    
    # Extract all event_ids for batch query
    event_ids = [e.get('event_id') for e in events_to_update if e.get('event_id')]
    
    if not event_ids:
        return events_to_create, events_to_update, []
    
    try:
        # BATCH QUERY: Get all events in one call (performance!)
        select_fields = 'id, category_id'
        existing_events = client.table('events') \
            .select(select_fields) \
            .in_('id', event_ids) \
            .eq('user_id', user_id) \
            .execute()
        
        # Build lookup map: event_id → category_id
        existing_map = {e['id']: e['category_id'] for e in existing_events.data}
        
    except Exception as e:
        # If batch query fails, default to treating all as CREATE (safer!)
        warnings.append(f"⚠️ Could not validate event IDs (error: {str(e)}). All marked for creation.")
        return events_to_create + events_to_update, [], warnings
    
    # Validate each event and reclassify if needed
    valid_updates = []
    reclassified_creates = []
    invalid_not_found = []
    invalid_category_mismatch = []
    
    for event_data in events_to_update:
        event_id = event_data.get('event_id')
        category_path = event_data.get('Category_Path', '')
        
        # Check 1: Does event exist?
        if event_id not in existing_map:
            # Event doesn't exist → CREATE
            event_data['_reclassified'] = True
            event_data['_reason'] = 'not_found'
            reclassified_creates.append(event_data)
            invalid_not_found.append(event_id[:8] + '...')
            continue
        
        # Check 2: Does category match?
        existing_category_id = existing_map[event_id]
        expected_category_id = cat_by_path.get(category_path)
        
        if existing_category_id != expected_category_id:
            # Wrong category → CREATE (prevents data corruption!)
            event_data['_reclassified'] = True
            event_data['_reason'] = 'category_mismatch'
            reclassified_creates.append(event_data)
            invalid_category_mismatch.append(event_id[:8] + '...')
            continue
        
        # All checks passed → Keep as UPDATE
        valid_updates.append(event_data)
    
    # Generate summary warnings (not per-record!)
    if reclassified_creates:
        total = len(reclassified_creates)
        not_found_count = len(invalid_not_found)
        mismatch_count = len(invalid_category_mismatch)
        
        warning_msg = f"⚠️ **{total} event(s) had invalid event IDs and were created as NEW events:**"
        
        if not_found_count > 0:
            warning_msg += f"\n  - {not_found_count} event ID(s) not found in database"
            if not_found_count <= 5:
                warning_msg += f" ({', '.join(invalid_not_found)})"
        
        if mismatch_count > 0:
            warning_msg += f"\n  - {mismatch_count} event ID(s) belonged to different categories"
            if mismatch_count <= 5:
                warning_msg += f" ({', '.join(invalid_category_mismatch)})"
        
        warning_msg += "\n\n💡 **Tip:** When adding new events in Excel, clear the event_id column (Col A) to avoid this."
        
        warnings.append(warning_msg)
    
    # Merge reclassified events into creates
    updated_creates = events_to_create + reclassified_creates
    
    return updated_creates, valid_updates, warnings


def validate_import_data(
    events_to_create: List[Dict],
    events_to_update: List[Dict],
    categories_dict: Dict[str, Dict],
    attribute_definitions: List[Dict]
) -> Tuple[List[Dict], List[Dict], List[str]]:
    """Validate import data and return validated events + errors."""
    errors = []
    valid_creates = []
    valid_updates = []
    
    cat_by_path = {info['full_path']: cat_id for cat_id, info in categories_dict.items()}
    
    attr_by_cat_name = {}
    for attr_def in attribute_definitions:
        key = (attr_def['category_id'], attr_def['name'])
        attr_by_cat_name[key] = attr_def
    
    for idx, event in enumerate(events_to_create, start=1):
        event_errors = []
        
        if not event.get('event_date'):
            event_errors.append(f"Row {idx}: event_date is required")
        
        cat_path = event.get('Category_Path', '')
        if not cat_path:
            event_errors.append(f"Row {idx}: Category_Path is required")
        elif cat_path not in cat_by_path:
            event_errors.append(f"Row {idx}: Category_Path '{cat_path}' not found")
        
        if event_errors:
            errors.extend(event_errors)
        else:
            event['_category_id'] = cat_by_path.get(cat_path)
            valid_creates.append(event)
    
    for idx, event in enumerate(events_to_update, start=1):
        event_errors = []
        
        event_id = event.get('event_id')
        if not event_id:
            event_errors.append(f"Update row {idx}: event_id is required")
        
        if not event.get('event_date'):
            event_errors.append(f"Update row {idx}: event_date is required")
        
        if event_errors:
            errors.extend(event_errors)
        else:
            valid_updates.append(event)
    
    return valid_creates, valid_updates, errors


def apply_import_changes(
    client, user_id: str,
    events_to_create: List[Dict],
    events_to_update: List[Dict],
    categories_dict: Dict[str, Dict],
    attribute_definitions: List[Dict]
) -> Tuple[int, int, List[str]]:
    """
    Apply import changes to database.
    
    V2.5.3 NEW: Multi-level event creation from single Excel row.
    - One Excel row can create MULTIPLE events (one per hierarchy level)
    - All events from same row share SAME session_start timestamp
    - Only creates event if at least ONE attribute for that level is populated
    - Only populated attributes are saved (empty attributes skipped)
    """
    created = 0
    updated = 0
    errors = []
    
    # Build mapping: (category_id, attr_name) → attr_definition
    attr_by_cat_name = {}
    for attr_def in attribute_definitions:
        key = (attr_def['category_id'], attr_def['name'])
        attr_by_cat_name[key] = attr_def
    
    # ─────────────────────────────────────────
    # CREATE EVENTS (V2.5.3: Multi-level support)
    # ─────────────────────────────────────────
    for event_data in events_to_create:
        try:
            # Extract Category_Path and get all hierarchy levels
            category_path = event_data.get('Category_Path', '')
            hierarchy_levels = get_hierarchy_levels_for_path(category_path, categories_dict)
            
            if not hierarchy_levels:
                errors.append(f"Invalid category path: {category_path}")
                continue
            
            # Parse date once (shared by all events from this row)
            event_date = event_data.get('event_date')
            if isinstance(event_date, datetime):
                event_date = event_date.date().isoformat()
            elif isinstance(event_date, date):
                event_date = event_date.isoformat()
            else:
                # V2.4.8: Flexible date parsing for string dates
                try:
                    from dateutil import parser
                    # dayfirst=True handles Croatian format (DD.MM.YYYY)
                    parsed_date = parser.parse(str(event_date), dayfirst=True)
                    event_date = parsed_date.date().isoformat()
                except ImportError:
                    # Fallback if dateutil not available
                    date_str = str(event_date).strip()
                    for fmt in ['%Y-%m-%d', '%d.%m.%Y', '%d.%M.%Y', '%d/%m/%Y']:
                        try:
                            parsed_date = datetime.strptime(date_str, fmt)
                            event_date = parsed_date.date().isoformat()
                            break
                        except:
                            continue
                    else:
                        event_date = str(event_date)
                except Exception:
                    event_date = str(event_date)
            
            # V2.5.3: Parse session_start once (SHARED by all events from this row!)
            session_start = event_data.get('session_start', '09:00')
            
            # Parse time to timestamp (default to 09:00 if empty or '09:00')
            try:
                # If empty/None, use default '09:00'
                if not session_start or str(session_start).strip() == '':
                    session_start = '09:00'
                
                # Parse HH:MM to timestamp
                time_parts = str(session_start).split(':')
                if len(time_parts) == 2:
                    hour, minute = int(time_parts[0]), int(time_parts[1])
                    # Create datetime from date + time
                    if isinstance(event_date, str):
                        date_obj = datetime.strptime(event_date, '%Y-%m-%d').date()
                    else:
                        date_obj = event_date
                    dt = datetime.combine(date_obj, datetime.min.time().replace(hour=hour, minute=minute))
                    session_start = dt.isoformat()
                else:
                    # Invalid format - use default 09:00
                    if isinstance(event_date, str):
                        date_obj = datetime.strptime(event_date, '%Y-%m-%d').date()
                    else:
                        date_obj = event_date
                    dt = datetime.combine(date_obj, datetime.min.time().replace(hour=9, minute=0))
                    session_start = dt.isoformat()
            except:
                # Fallback to 09:00 on any error
                try:
                    if isinstance(event_date, str):
                        date_obj = datetime.strptime(event_date, '%Y-%m-%d').date()
                    else:
                        date_obj = event_date
                    dt = datetime.combine(date_obj, datetime.min.time().replace(hour=9, minute=0))
                    session_start = dt.isoformat()
                except:
                    session_start = None
            
            # V2.5.3: Create event for EACH hierarchy level that has populated attributes
            for partial_path, category_id in hierarchy_levels:
                # Find which attributes belong to this category level
                level_attributes = {}
                
                for key, value in event_data.items():
                    if key in FIXED_COLUMNS or key.startswith('_') or not value:
                        continue
                    
                    # Check if this attribute belongs to current category level
                    attr_def = attr_by_cat_name.get((category_id, key))
                    if attr_def:
                        level_attributes[key] = (value, attr_def)
                
                # Only create event if at least ONE attribute is populated for this level
                if not level_attributes:
                    continue
                
                # Create event for this hierarchy level
                new_event = {
                    'user_id': user_id,
                    'category_id': category_id,
                    'event_date': event_date,
                    'session_start': session_start,  # SAME timestamp for all levels!
                    'comment': event_data.get('comment', '') or None
                }
                
                result = client.table('events').insert(new_event).execute()
                event_id = result.data[0]['id']
                
                # Insert only populated attributes for this level
                for attr_name, (value, attr_def) in level_attributes.items():
                    attr_data = {
                        'event_id': event_id,
                        'attribute_definition_id': attr_def['id'],
                        'user_id': user_id,
                        'value_text': None,
                        'value_number': None,
                        'value_datetime': None,
                        'value_boolean': None
                    }
                    
                    data_type = attr_def.get('data_type', 'text')
                    if data_type == 'number':
                        attr_data['value_number'] = float(value) if value else None
                    elif data_type == 'boolean':
                        attr_data['value_boolean'] = bool(value)
                    elif data_type == 'datetime':
                        attr_data['value_datetime'] = str(value)
                    else:
                        attr_data['value_text'] = str(value)
                    
                    client.table('event_attributes').insert(attr_data).execute()
                
                created += 1
            
        except Exception as e:
            errors.append(f"Error creating event: {str(e)}")
    
    # ─────────────────────────────────────────
    # UPDATE EVENTS (V2.5.4: Multi-level support added)
    # ─────────────────────────────────────────
    for event_data in events_to_update:
        try:
            event_id = event_data.get('event_id')
            
            # First, fetch existing event to get its category_id
            select_fields = 'id, category_id, event_attributes(id, attribute_definition_id)'
            existing = client.table('events') \
                .select(select_fields) \
                .eq('id', event_id) \
                .eq('user_id', user_id) \
                .single() \
                .execute()
            
            if not existing.data:
                errors.append(f"Event {event_id} not found")
                continue
            
            existing_category_id = existing.data.get('category_id')
            
            # Extract hierarchy levels from Category_Path
            category_path = event_data.get('Category_Path', '')
            hierarchy_levels = get_hierarchy_levels_for_path(category_path, categories_dict)
            
            if not hierarchy_levels:
                errors.append(f"Invalid category path for event {event_id}: {category_path}")
                continue
            
            # Parse date once (shared by all events from this row)
            event_date = event_data.get('event_date')
            if isinstance(event_date, datetime):
                event_date = event_date.date().isoformat()
            elif isinstance(event_date, date):
                event_date = event_date.isoformat()
            else:
                # V2.4.8: Flexible date parsing for string dates
                try:
                    from dateutil import parser
                    parsed_date = parser.parse(str(event_date), dayfirst=True)
                    event_date = parsed_date.date().isoformat()
                except ImportError:
                    date_str = str(event_date).strip()
                    for fmt in ['%Y-%m-%d', '%d.%m.%Y', '%d.%M.%Y', '%d/%m/%Y']:
                        try:
                            parsed_date = datetime.strptime(date_str, fmt)
                            event_date = parsed_date.date().isoformat()
                            break
                        except:
                            continue
                    else:
                        event_date = str(event_date)
                except Exception:
                    event_date = str(event_date)
            
            # V2.5.3: Parse session_start once (SHARED by all events from this row!)
            session_start = event_data.get('session_start', '09:00')
            
            # Parse time to timestamp (default to 09:00 if empty or '09:00')
            try:
                # If empty/None, use default '09:00'
                if not session_start or str(session_start).strip() == '':
                    session_start = '09:00'
                
                # Parse HH:MM to timestamp
                time_parts = str(session_start).split(':')
                if len(time_parts) == 2:
                    hour, minute = int(time_parts[0]), int(time_parts[1])
                    # Create datetime from date + time
                    if isinstance(event_date, str):
                        date_obj = datetime.strptime(event_date, '%Y-%m-%d').date()
                    else:
                        date_obj = event_date
                    dt = datetime.combine(date_obj, datetime.min.time().replace(hour=hour, minute=minute))
                    session_start = dt.isoformat()
                else:
                    # Invalid format - use default 09:00
                    if isinstance(event_date, str):
                        date_obj = datetime.strptime(event_date, '%Y-%m-%d').date()
                    else:
                        date_obj = event_date
                    dt = datetime.combine(date_obj, datetime.min.time().replace(hour=9, minute=0))
                    session_start = dt.isoformat()
            except:
                # Fallback to 09:00 on any error
                try:
                    if isinstance(event_date, str):
                        date_obj = datetime.strptime(event_date, '%Y-%m-%d').date()
                    else:
                        date_obj = event_date
                    dt = datetime.combine(date_obj, datetime.min.time().replace(hour=9, minute=0))
                    session_start = dt.isoformat()
                except:
                    session_start = None
            
            # V2.5.4 NEW: Create parent events for levels with populated attributes
            # (only for parent levels, NOT for the existing event's category)
            for partial_path, category_id in hierarchy_levels:
                # Skip the existing event's category (we'll UPDATE it later)
                if category_id == existing_category_id:
                    continue
                
                # Find which attributes belong to this parent category level
                level_attributes = {}
                
                for key, value in event_data.items():
                    if key in FIXED_COLUMNS or key.startswith('_') or not value:
                        continue
                    
                    # Check if this attribute belongs to current parent category level
                    attr_def = attr_by_cat_name.get((category_id, key))
                    if attr_def:
                        level_attributes[key] = (value, attr_def)
                
                # Only create parent event if at least ONE attribute is populated for this level
                if not level_attributes:
                    continue
                
                # Create parent event with SAME timestamp!
                new_parent_event = {
                    'user_id': user_id,
                    'category_id': category_id,
                    'event_date': event_date,
                    'session_start': session_start,  # SAME timestamp as child!
                    'comment': event_data.get('comment', '') or None
                }
                
                result = client.table('events').insert(new_parent_event).execute()
                parent_event_id = result.data[0]['id']
                
                # Insert only populated attributes for this parent level
                for attr_name, (value, attr_def) in level_attributes.items():
                    attr_data = {
                        'event_id': parent_event_id,
                        'attribute_definition_id': attr_def['id'],
                        'user_id': user_id,
                        'value_text': None,
                        'value_number': None,
                        'value_datetime': None,
                        'value_boolean': None
                    }
                    
                    data_type = attr_def.get('data_type', 'text')
                    if data_type == 'number':
                        attr_data['value_number'] = float(value) if value else None
                    elif data_type == 'boolean':
                        attr_data['value_boolean'] = bool(value)
                    elif data_type == 'datetime':
                        attr_data['value_datetime'] = str(value)
                    else:
                        attr_data['value_text'] = str(value)
                    
                    client.table('event_attributes').insert(attr_data).execute()
                
                created += 1  # Count parent event as created
            
            # Now UPDATE the existing child event (normal UPDATE logic)
            updates = {
                'event_date': event_date,
                'session_start': session_start,
                'comment': event_data.get('comment', '') or None,
                'edited_at': datetime.now().isoformat()
            }
            
            client.table('events') \
                .update(updates) \
                .eq('id', event_id) \
                .eq('user_id', user_id) \
                .execute()
            
            existing_attrs = {
                ea['attribute_definition_id']: ea['id']
                for ea in existing.data.get('event_attributes', [])
            }
            
            # Update/insert attributes for the child event (only its own category's attributes)
            for key, value in event_data.items():
                if key in FIXED_COLUMNS or key.startswith('_'):
                    continue
                
                # Only update attributes that belong to the existing event's category
                attr_def = attr_by_cat_name.get((existing_category_id, key))
                if not attr_def:
                    continue
                
                attr_def_id = attr_def['id']
                data_type = attr_def.get('data_type', 'text')
                
                attr_update = {
                    'value_text': None,
                    'value_number': None,
                    'value_datetime': None,
                    'value_boolean': None
                }
                
                if value is not None and value != '':
                    if data_type == 'number':
                        attr_update['value_number'] = float(value)
                    elif data_type == 'boolean':
                        attr_update['value_boolean'] = bool(value)
                    elif data_type == 'datetime':
                        attr_update['value_datetime'] = str(value)
                    else:
                        attr_update['value_text'] = str(value)
                
                if attr_def_id in existing_attrs:
                    client.table('event_attributes') \
                        .update(attr_update) \
                        .eq('id', existing_attrs[attr_def_id]) \
                        .eq('user_id', user_id) \
                        .execute()
                elif value is not None and value != '':
                    attr_update['event_id'] = event_id
                    attr_update['attribute_definition_id'] = attr_def_id
                    attr_update['user_id'] = user_id
                    client.table('event_attributes').insert(attr_update).execute()
            
            updated += 1
            
        except Exception as e:
            errors.append(f"Error updating event {event_data.get('event_id')}: {str(e)}")
    
    return created, updated, errors


# ============================================
# HIGH-LEVEL EXPORT FUNCTION
# ============================================

def export_events_to_excel(
    client, user_id: str,
    area_id: Optional[str] = None,
    category_ids: Optional[List[str]] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    sort_order: str = 'desc'  # V2.5.5: User-selected sort order
) -> Tuple[bytes, int, str]:
    """
    High-level function to export events to Excel.
    
    V2.5.5 NEW:
    - Session-based merging: Events with same timestamp in parent-child hierarchy
      are merged into single row (leaf event with all attributes)
    - Respects UI sort order (newest/oldest first)
    
    V2.5.1: Fixed filter logic - include only selected category branch + descendants
    V2.5.0: Includes parent category attributes in export.
    """
    try:
        categories_dict = load_categories_dict(client, user_id)
        
        # Determine effective category_ids
        effective_category_ids = category_ids
        
        if not effective_category_ids and area_id:
            # Get all categories for this area (no specific category selected)
            effective_category_ids = get_category_ids_for_area(client, user_id, area_id)
            if not effective_category_ids:
                return b'', 0, "No categories found for selected area"
        
        if not effective_category_ids:
            # No filter - get all categories
            effective_category_ids = list(categories_dict.keys())
        
        # V2.5.1 FIX: If category_ids were provided (user selected specific category),
        # expand to include ALL descendants (e.g., Cardio → Cardio + Running + Cycling + Swimming)
        # This ensures we don't mix Cardio and Strength branches
        if category_ids:  # Only expand if explicitly provided categories
            effective_category_ids = get_all_descendant_category_ids(
                categories_dict, effective_category_ids
            )
        
        # V2.5.0: Include parent categories to show their attributes too
        category_ids_with_parents = get_category_ids_with_parents(
            categories_dict, effective_category_ids
        )
        
        attribute_definitions = load_attribute_definitions_for_categories(
            client, user_id, category_ids_with_parents
        )
        
        events = load_events_for_export(
            client, user_id,
            category_ids=effective_category_ids,  # Still filter events by original selection
            date_from=date_from,
            date_to=date_to,
            sort_order=sort_order  # V2.5.5: Pass user sort order
        )
        
        if not events:
            return b'', 0, "No events found matching filters"
        
        # V2.5.5 NEW: Merge hierarchical session events
        # This reduces export rows by combining parent+child events with same timestamp
        merged_events = merge_session_events(events, categories_dict)
        
        excel_bytes = create_events_excel_v2(merged_events, attribute_definitions, categories_dict)
        
        # Return count of merged events (what user sees in Excel)
        return excel_bytes, len(merged_events), ""
        
    except Exception as e:
        return b'', 0, f"Export error: {str(e)}"


# ============================================
# HIGH-LEVEL IMPORT FUNCTION
# ============================================

def import_events_from_excel(
    client, user_id: str, file_bytes: bytes
) -> Tuple[int, int, List[str]]:
    """
    High-level function to import events from Excel.
    
    V2.4.6 CHANGES:
    - Added smart_reclassify_events for auto-reclassification
    - Auto-converts invalid event_ids to CREATE (no manual cleanup!)
    - Returns warnings instead of hard errors for reclassified events
    """
    events_to_create, events_to_update, parse_error = parse_events_excel(file_bytes)
    
    if parse_error:
        return 0, 0, [parse_error]
    
    if not events_to_create and not events_to_update:
        return 0, 0, ["No events found in file"]
    
    categories_dict = load_categories_dict(client, user_id)
    all_category_ids = list(categories_dict.keys())
    attribute_definitions = load_attribute_definitions_for_categories(
        client, user_id, all_category_ids
    )
    
    # V2.4.6 NEW: Smart validation and reclassification
    events_to_create, events_to_update, reclassify_warnings = smart_reclassify_events(
        client, user_id, events_to_create, events_to_update, categories_dict
    )
    
    valid_creates, valid_updates, validation_errors = validate_import_data(
        events_to_create, events_to_update, categories_dict, attribute_definitions
    )
    
    if validation_errors:
        return 0, 0, validation_errors
    
    created, updated, apply_errors = apply_import_changes(
        client, user_id, valid_creates, valid_updates,
        categories_dict, attribute_definitions
    )
    
    # Combine warnings with any apply errors
    all_messages = reclassify_warnings + apply_errors
    
    return created, updated, all_messages
