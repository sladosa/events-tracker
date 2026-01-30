"""Enhanced Structure Exporter v5 - WITH DEPENDENCIES

CHANGES FROM V4 (2026-01-30):
- Added DependsOn column (P) - GREEN color
- Added WhenValue column (Q) - GREEN color
- Renamed ValidationMin to TextOptions (M) for clarity
- Support for exporting dependency configurations from validation_rules
- Backward compatible: existing files without dependencies still work

Column layout (HierarchicalView):
A Type (pink)
B Level (pink)
C SortOrder (yellow)
D Area (pink)
E CategoryPath (yellow)
F Category (blue)
G AttributeName (blue)
H DataType (blue)
I Unit (blue)
J IsRequired (blue)
K ValidationType (blue)
L DefaultValue (blue)
M TextOptions (blue) - was ValidationMin
N ValidationMax (blue)
O Description (blue)
P DependsOn (green) - NEW
Q WhenValue (green) - NEW

Compatible with HierarchicalParser v5.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.worksheet.properties import Outline
from datetime import datetime
from typing import Dict, List, Optional, Tuple
import json


class EnhancedStructureExporter:
    def __init__(self, client, user_id: str, filter_area: Optional[str] = None, filter_category: Optional[str] = None):
        self.client = client
        self.user_id = user_id
        self.filter_area = filter_area if filter_area != 'All Areas' else None
        self.filter_category = filter_category if filter_category != 'All Categories' else None

        self.PINK_FILL = PatternFill(start_color='FFE6F0', end_color='FFE6F0', fill_type='solid')
        self.BLUE_FILL = PatternFill(start_color='E6F2FF', end_color='E6F2FF', fill_type='solid')
        self.YELLOW_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        self.GREEN_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')  # NEW
        self.HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.BOLD_FONT = Font(bold=True, color='FFFFFF')
        self.THIN_BORDER = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
        )
        self.CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
        self.LEFT_ALIGN = Alignment(horizontal='left', vertical='center')

        self.HEADER_COMMENTS = {
            'A': 'Type: Area, Category, or Attribute. Do NOT change for existing rows.',
            'B': 'Level: Auto-calculated from CategoryPath depth. Read-only.',
            'C': 'SortOrder: Position within parent element. Edit to change display order.',
            'D': 'Area: Auto-extracted from CategoryPath. Read-only.',
            'E': 'CategoryPath: KEY identifier. For existing rows DO NOT change.',
            'F': 'Category: Must match the LAST part of CategoryPath.',
            'G': 'AttributeName: Only for Attribute rows.',
            'H': 'DataType: number, text, datetime, boolean, link, image.',
            'I': 'Unit: kg, hours, EUR, km, bpm...',
            'J': 'IsRequired: TRUE or FALSE.',
            'K': 'ValidationType: suggest (default for text), enum, none.',
            'L': 'DefaultValue: Default value for new events.',
            'M': 'TextOptions: Pipe-separated options (Run|Walk|Cycle). For numbers: minimum value.',
            'N': 'ValidationMax: Maximum allowed value (number only).',
            'O': 'Description: Documentation / notes.',
            'P': 'DependsOn: Slug of parent attribute (for dependent dropdowns). Must be in same category.',
            'Q': 'WhenValue: Value of parent attribute for these options. Use "*" for fallback/default.',
        }

    def export_hierarchical_view(self, output_path: Optional[str] = None) -> str:
        rows = self._load_hierarchical_data()
        wb = Workbook()
        ws = wb.active
        ws.title = 'HierarchicalView'

        self._setup_headers(ws)
        self._populate_data(ws, rows)
        self._add_data_validations(ws)
        self._setup_column_groups(ws)
        self._autosize_columns(ws)

        ws.freeze_panes = 'G3'
        ws.auto_filter.ref = f'A2:Q{len(rows)+2}'

        self._add_help_sheet(wb)

        if not output_path:
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = f'structure_hierarchical_v5_{ts}.xlsx'
        wb.save(output_path)
        return output_path

    def _load_hierarchical_data(self) -> List[Dict]:
        """Load hierarchical structure from database."""
        rows: List[Dict] = []
        
        areas_q = self.client.table('areas').select('*').eq('user_id', self.user_id)
        if self.filter_area:
            areas_q = areas_q.eq('name', self.filter_area)
        areas = areas_q.order('sort_order').execute().data or []

        for area in areas:
            rows.append(self._create_row(
                row_type='Area',
                level=0,
                sort_order=area.get('sort_order', 0),
                area_name=area.get('name', ''),
                category_path=area.get('name', ''),
                description=area.get('description', '')
            ))
            self._load_categories_recursive(area['id'], area.get('name', ''), rows)

        if self.filter_category and rows:
            rows = [r for r in rows if 
                    self.filter_category in r.get('CategoryPath', '') or 
                    r.get('Type') == 'Area']
        
        return rows

    def _create_row(self, row_type: str, level: int, sort_order: int, area_name: str,
                    category_path: str, category_name: str = '', attribute_name: str = '',
                    data_type: str = '', unit: str = '', is_required: str = '',
                    validation_type: str = '', default_value: str = '', text_options: str = '',
                    validation_max: str = '', description: str = '', 
                    depends_on: str = '', when_value: str = '') -> Dict:
        """Create a standardized row dictionary."""
        return {
            'Type': row_type,
            'Level': level,
            'SortOrder': sort_order,
            'Area': area_name,
            'CategoryPath': category_path,
            'Category': category_name,
            'AttributeName': attribute_name,
            'DataType': data_type,
            'Unit': unit,
            'IsRequired': is_required,
            'ValidationType': validation_type,
            'DefaultValue': default_value,
            'TextOptions': text_options,
            'ValidationMax': validation_max,
            'Description': description,
            'DependsOn': depends_on,
            'WhenValue': when_value,
        }

    def _load_categories_recursive(self, area_id: str, area_name: str, rows: List[Dict], 
                                    parent_id: Optional[str] = None, parent_path: str = '', level: int = 1):
        """Recursively load categories and their attributes."""
        q = self.client.table('categories').select('*').eq('user_id', self.user_id).eq('area_id', area_id).eq('level', level)
        if parent_id:
            q = q.eq('parent_category_id', parent_id)
        else:
            q = q.is_('parent_category_id', 'null')
        cats = q.order('sort_order').execute().data or []

        for cat in cats:
            catpath = f"{parent_path} > {cat['name']}" if parent_path else f"{area_name} > {cat['name']}"
            rows.append(self._create_row(
                row_type='Category',
                level=level,
                sort_order=cat.get('sort_order', 0),
                area_name=area_name,
                category_path=catpath,
                category_name=cat.get('name', ''),
                description=cat.get('description', '')
            ))

            # Load attributes for this category
            attrs = self.client.table('attribute_definitions').select('*').eq('user_id', self.user_id).eq('category_id', cat['id']).order('sort_order').execute().data or []
            
            for attr in attrs:
                attr_rows = self._process_attribute(attr, area_name, catpath, cat.get('name', ''), level)
                rows.extend(attr_rows)

            if level < 10:
                self._load_categories_recursive(area_id, area_name, rows, cat['id'], catpath, level + 1)

    def _process_attribute(self, attr: Dict, area_name: str, catpath: str, 
                           category_name: str, level: int) -> List[Dict]:
        """Process an attribute and return one or more rows (multiple if has dependencies)."""
        vr = self._parse_validation_rules(attr.get('validation_rules'))
        datatype = (attr.get('data_type') or 'text').strip()
        vtype = (vr.get('type') or '').strip()

        if not vtype and datatype == 'text':
            vtype = 'suggest'

        # Check for dependencies
        depends_on_config = vr.get('depends_on', {})
        
        if depends_on_config and 'options_map' in depends_on_config:
            # Attribute has dependencies - create multiple rows
            return self._create_dependency_rows(
                attr, area_name, catpath, category_name, level, 
                datatype, vtype, depends_on_config
            )
        else:
            # Simple attribute - single row
            text_options = self._extract_text_options(vr, datatype)
            valmin = vr.get('min', '')
            valmax = vr.get('max', '')
            
            return [self._create_row(
                row_type='Attribute',
                level=level + 1,
                sort_order=attr.get('sort_order', 0),
                area_name=area_name,
                category_path=catpath,
                category_name=category_name,
                attribute_name=attr.get('name', ''),
                data_type=datatype,
                unit=attr.get('unit', ''),
                is_required='TRUE' if attr.get('is_required', False) else 'FALSE',
                validation_type=vtype,
                default_value=attr.get('default_value', ''),
                text_options=text_options if datatype == 'text' else str(valmin) if valmin else '',
                validation_max='' if datatype == 'text' else str(valmax) if valmax else '',
                description=attr.get('description', ''),
            )]

    def _create_dependency_rows(self, attr: Dict, area_name: str, catpath: str,
                                 category_name: str, level: int, datatype: str,
                                 vtype: str, depends_on_config: Dict) -> List[Dict]:
        """Create multiple rows for an attribute with dependencies."""
        rows = []
        parent_slug = depends_on_config.get('attribute_slug', '')
        options_map = depends_on_config.get('options_map', {})
        
        for when_value, options_list in options_map.items():
            if isinstance(options_list, list):
                text_options = '|'.join([str(x) for x in options_list if str(x).strip()])
            else:
                text_options = str(options_list) if options_list else ''
            
            rows.append(self._create_row(
                row_type='Attribute',
                level=level + 1,
                sort_order=attr.get('sort_order', 0),
                area_name=area_name,
                category_path=catpath,
                category_name=category_name,
                attribute_name=attr.get('name', ''),
                data_type=datatype,
                unit=attr.get('unit', ''),
                is_required='TRUE' if attr.get('is_required', False) else 'FALSE',
                validation_type=vtype,
                default_value=attr.get('default_value', ''),
                text_options=text_options,
                validation_max='',
                description=attr.get('description', ''),
                depends_on=parent_slug,
                when_value=when_value,
            ))
        
        # Add fallback row if not present
        if '*' not in options_map:
            rows.append(self._create_row(
                row_type='Attribute',
                level=level + 1,
                sort_order=attr.get('sort_order', 0),
                area_name=area_name,
                category_path=catpath,
                category_name=category_name,
                attribute_name=attr.get('name', ''),
                data_type=datatype,
                unit=attr.get('unit', ''),
                is_required='TRUE' if attr.get('is_required', False) else 'FALSE',
                validation_type=vtype,
                default_value=attr.get('default_value', ''),
                text_options='',
                validation_max='',
                description=attr.get('description', ''),
                depends_on=parent_slug,
                when_value='*',
            ))
        
        return rows

    def _parse_validation_rules(self, vr) -> Dict:
        """Parse validation_rules, handling double-escaped JSON."""
        if isinstance(vr, str):
            try:
                vr = json.loads(vr)
                if isinstance(vr, str):
                    vr = json.loads(vr)
            except Exception:
                vr = {}
        if not isinstance(vr, dict):
            vr = {}
        return vr

    def _extract_text_options(self, vr: Dict, datatype: str) -> str:
        """Extract text options from validation rules."""
        if datatype != 'text':
            return ''
        opt_list = vr.get('enum') or vr.get('suggest') or vr.get('static_options')
        if isinstance(opt_list, list):
            return '|'.join([str(x) for x in opt_list if str(x).strip()])
        return ''

    def _setup_headers(self, ws):
        headers = [
            'Type', 'Level', 'SortOrder', 'Area', 'CategoryPath', 'Category', 'AttributeName', 
            'DataType', 'Unit', 'IsRequired', 'ValidationType', 'DefaultValue', 'TextOptions', 
            'ValidationMax', 'Description', 'DependsOn', 'WhenValue'
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = self.BOLD_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.CENTER_ALIGN
            cell.border = self.THIN_BORDER

        col_letters = 'ABCDEFGHIJKLMNOPQ'
        for i, col in enumerate(col_letters, start=1):
            if col in self.HEADER_COMMENTS:
                c = ws.cell(row=2, column=i)
                comment = Comment(self.HEADER_COMMENTS[col], 'Events Tracker')
                comment.width = 380
                comment.height = 150
                c.comment = comment

    def _populate_data(self, ws, rows: List[Dict]):
        col_order = [
            'Type', 'Level', 'SortOrder', 'Area', 'CategoryPath', 'Category', 'AttributeName',
            'DataType', 'Unit', 'IsRequired', 'ValidationType', 'DefaultValue', 'TextOptions',
            'ValidationMax', 'Description', 'DependsOn', 'WhenValue'
        ]
        
        for r, row_data in enumerate(rows, start=3):
            for c, col_name in enumerate(col_order, start=1):
                val = row_data.get(col_name, '')
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = self.THIN_BORDER
                
                # Apply colors
                if c in (1, 2, 4):  # A, B, D - pink
                    cell.fill = self.PINK_FILL
                    cell.alignment = self.CENTER_ALIGN if c != 4 else self.LEFT_ALIGN
                elif c in (3, 5):  # C, E - yellow
                    cell.fill = self.YELLOW_FILL
                    cell.alignment = self.CENTER_ALIGN if c == 3 else self.LEFT_ALIGN
                elif c in (16, 17):  # P, Q - green (NEW)
                    cell.fill = self.GREEN_FILL
                    cell.alignment = self.LEFT_ALIGN
                else:  # blue
                    cell.fill = self.BLUE_FILL
                    cell.alignment = self.LEFT_ALIGN if c in (6, 7, 12, 13, 15) else self.CENTER_ALIGN

            # Area formula
            area_formula = f'=IFERROR(LEFT(E{r},FIND(" > ",E{r})-1),E{r})'
            ws.cell(r, 4).value = area_formula
            ws.cell(r, 4).fill = self.PINK_FILL

    def _add_data_validations(self, ws):
        maxrow = max(ws.max_row, 100)

        type_dv = DataValidation(type='list', formula1='"Area,Category,Attribute"', allow_blank=False)
        ws.add_data_validation(type_dv)
        type_dv.add(f'A3:A{maxrow}')

        datatype_dv = DataValidation(type='list', formula1='"number,text,datetime,boolean,link,image"', allow_blank=True)
        ws.add_data_validation(datatype_dv)
        datatype_dv.add(f'H3:H{maxrow}')

        required_dv = DataValidation(type='list', formula1='"TRUE,FALSE"', allow_blank=True)
        ws.add_data_validation(required_dv)
        required_dv.add(f'J3:J{maxrow}')

        vtype_dv = DataValidation(type='list', formula1='"suggest,enum,none"', allow_blank=True)
        ws.add_data_validation(vtype_dv)
        vtype_dv.add(f'K3:K{maxrow}')

    def _setup_column_groups(self, ws):
        ws.sheet_properties.outlinePr = Outline(summaryRight=False)
        ws.column_dimensions['B'].outlineLevel = 1
        ws.column_dimensions['D'].outlineLevel = 1
        ws.column_dimensions['F'].outlineLevel = 1
        for col in ['I', 'J', 'K', 'L', 'M', 'N']:
            ws.column_dimensions[col].outlineLevel = 1

    def _autosize_columns(self, ws):
        col_widths = {
            'A': 10, 'B': 8, 'C': 10, 'D': 10, 'E': 50, 'F': 12, 'G': 18,
            'H': 10, 'I': 8, 'J': 10, 'K': 14, 'L': 12, 'M': 45,
            'N': 12, 'O': 25, 'P': 15, 'Q': 12
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

    def _add_help_sheet(self, wb: Workbook):
        ws = wb.create_sheet('Help')
        
        title = 'EVENTS TRACKER - Structure Import/Export Guide V5'
        ws.cell(1, 1, title).font = Font(bold=True, size=14, color='FFFFFF')
        ws.cell(1, 1).fill = self.HEADER_FILL
        
        lines = [
            '',
            '═══════════════════════════════════════════════════════════════════',
            '📋 COLUMN REFERENCE (V5 - with Dependencies)',
            '═══════════════════════════════════════════════════════════════════',
            '',
            '🟪 PINK COLUMNS (Auto-calculated / Read-only):',
            '   A - Type: Area, Category, or Attribute',
            '   B - Level: Auto-calculated from CategoryPath depth',
            '   D - Area: Auto-extracted from CategoryPath',
            '',
            '🟨 YELLOW COLUMNS (Key identifiers - Edit ONLY for NEW rows):',
            '   C - SortOrder: Position within parent',
            '   E - CategoryPath: KEY identifier using " > " separator',
            '       ⚠️ For EXISTING rows DO NOT CHANGE - creates duplicates!',
            '',
            '🟦 BLUE COLUMNS (Freely editable):',
            '   F  - Category: Must match the LAST part of CategoryPath',
            '   G  - AttributeName: Name of the attribute',
            '   H  - DataType: number, text, datetime, boolean, link, image',
            '   I  - Unit: Measurement unit (kg, min, EUR, km...)',
            '   J  - IsRequired: TRUE or FALSE',
            '   K  - ValidationType: suggest (default), enum, none',
            '   L  - DefaultValue: Default value for new events',
            '   M  - TextOptions: Pipe-separated options (Run|Walk|Cycle)',
            '        For numbers: minimum value',
            '   N  - ValidationMax: Maximum value (number only)',
            '   O  - Description: Documentation / notes',
            '',
            '🟢 GREEN COLUMNS (NEW - Conditional Dropdowns):',
            '   P  - DependsOn: Slug of parent attribute that controls this dropdown',
            '        Must be in the SAME category',
            '   Q  - WhenValue: Value of parent attribute for these options',
            '        • Specific value: "Upp", "Low", "Core"',
            '        • Wildcard "*": Fallback for undefined parent values',
            '',
            '═══════════════════════════════════════════════════════════════════',
            '📌 HOW TO USE DEPENDENCIES (Conditional Dropdowns)',
            '═══════════════════════════════════════════════════════════════════',
            '',
            'Use case: exercise_name options depend on strength_type selection',
            '',
            'Step 1: Define the parent attribute (no DependsOn)',
            '| AttributeName  | TextOptions    | DependsOn | WhenValue |',
            '| strength_type  | Upp|Low|Core   |           |           |',
            '',
            'Step 2: Add child rows for EACH parent value',
            '| AttributeName  | TextOptions              | DependsOn      | WhenValue |',
            '| exercise_name  | pull.m|biceps|triceps    | strength_type  | Upp       |',
            '| exercise_name  | squat-bw|iskoraci        | strength_type  | Low       |',
            '| exercise_name  | plank|leg.raises         | strength_type  | Core      |',
            '| exercise_name  |                          | strength_type  | *         |',
            '',
            'Rules:',
            '• Same AttributeName can appear MULTIPLE TIMES with different WhenValue',
            '• DependsOn must reference attribute in SAME category',
            '• WhenValue = "*" is fallback (empty TextOptions = free text input)',
            '• All rows for same attribute should have same SortOrder',
            '• Import process merges all rows into single validation_rules JSON',
            '',
            '═══════════════════════════════════════════════════════════════════',
            '🎨 COLOR CODING (4 Colors)',
            '═══════════════════════════════════════════════════════════════════',
            '',
            '🟪 PINK = Auto-calculated, READ-ONLY',
            '🟨 YELLOW = KEY IDENTIFIER - Edit carefully',
            '🟦 BLUE = Freely EDITABLE',
            '🟢 GREEN = DEPENDENCIES (NEW in V5)',
            '',
            '═══════════════════════════════════════════════════════════════════',
            '✅ VALIDATION TYPES',
            '═══════════════════════════════════════════════════════════════════',
            '',
            'For TEXT attributes:',
            '• suggest (default): Free text + optional suggestions from TextOptions',
            '  User CAN type values not in the list',
            '  User CAN add new values through "Other..." option',
            '',
            '• enum: Strict dropdown - ONLY listed values allowed',
            '  User must pick from TextOptions list',
            '',
            '• none: No validation, no suggestions - pure free text',
            '',
            'For NUMBER attributes:',
            '  TextOptions (M) = minimum value',
            '  ValidationMax (N) = maximum value',
            '',
        ]
        
        for i, t in enumerate(lines, start=2):
            ws.cell(i, 1, t)
        
        ws.freeze_panes = 'A2'
        ws.column_dimensions['A'].width = 85
