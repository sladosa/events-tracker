"""
Error Reporter - Generate Excel with highlighted validation errors
====================================================================

Features:
- Open original Excel file
- Highlight cells with validation errors in yellow
- Add comments explaining each error
- Support for row-level and column-level errors
- Generate downloadable error report Excel

Dependencies: openpyxl
Last Modified: 2025-11-24 12:00 UTC
"""

import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.comments import Comment
from typing import List, Dict, Optional
import os
import tempfile


class ErrorReporter:
    """
    Generate Excel file with highlighted validation errors.
    """
    
    # Yellow fill for error cells
    ERROR_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ERROR_FONT = Font(color="FF0000", bold=True)  # Red bold text
    
    def __init__(self, excel_path: str, validation_errors: List):
        """
        Initialize ErrorReporter.
        
        Args:
            excel_path: Path to original Excel file
            validation_errors: List of ValidationError objects
        """
        self.excel_path = excel_path
        self.validation_errors = validation_errors
        self.wb = None
        self.ws = None
        self.column_map: Dict[str, int] = {}  # Map column names to column numbers
    
    def generate_error_report(self) -> str:
        """
        Generate Excel file with highlighted errors.
        
        Returns:
            Path to generated error report Excel file
        """
        # Load workbook
        self.wb = openpyxl.load_workbook(self.excel_path)
        self.ws = self.wb['Hierarchical_View']
        
        # Build column map (row 2 contains headers)
        self._build_column_map()
        
        # Process each validation error
        for error in self.validation_errors:
            self._highlight_error(error)
        
        # Save to temporary file
        output_path = self._get_output_path()
        self.wb.save(output_path)
        
        return output_path
    
    def _build_column_map(self):
        """
        Build mapping of column names to column numbers.
        Headers are in row 2.
        """
        header_row = 2
        for col_idx in range(1, self.ws.max_column + 1):
            cell = self.ws.cell(row=header_row, column=col_idx)
            if cell.value:
                col_name = str(cell.value).strip()
                self.column_map[col_name] = col_idx
    
    def _highlight_error(self, error):
        """
        Highlight a specific error in the Excel file.
        
        Args:
            error: ValidationError object with row, column, message
        """
        # Handle different error types
        if error.row == 0:
            # File-level or column-level error
            if error.column == "Columns":
                # Highlight entire header row in yellow
                self._highlight_header_error(error)
            elif error.column == "File" or error.column == "Database":
                # Add a comment in cell A1
                self._add_file_error(error)
        else:
            # Row-specific error
            self._highlight_cell_error(error)
    
    def _highlight_header_error(self, error):
        """
        Highlight header row for missing columns errors.
        
        Args:
            error: ValidationError object
        """
        header_row = 2
        
        # Highlight entire header row
        for col_idx in range(1, self.ws.max_column + 1):
            cell = self.ws.cell(row=header_row, column=col_idx)
            cell.fill = self.ERROR_FILL
        
        # Add comment to first cell with error message
        first_cell = self.ws.cell(row=header_row, column=1)
        comment = Comment(f"❌ ERROR:\n{error.message}", "Validation System")
        comment.width = 300
        comment.height = 100
        first_cell.comment = comment
    
    def _add_file_error(self, error):
        """
        Add file-level error as comment in cell A1.
        
        Args:
            error: ValidationError object
        """
        cell = self.ws.cell(row=1, column=1)
        cell.fill = self.ERROR_FILL
        cell.value = "⚠️ FILE ERROR"
        cell.font = self.ERROR_FONT
        
        comment = Comment(f"❌ FILE ERROR:\n{error.message}", "Validation System")
        comment.width = 300
        comment.height = 100
        cell.comment = comment
    
    def _highlight_cell_error(self, error):
        """
        Highlight specific cell with validation error.
        
        Args:
            error: ValidationError object with row and column
        """
        # Get Excel row (error.row is already Excel row number)
        excel_row = error.row
        
        # Get column number from column name
        col_num = self.column_map.get(error.column)
        
        if col_num:
            # Highlight the specific cell
            cell = self.ws.cell(row=excel_row, column=col_num)
            cell.fill = self.ERROR_FILL
            
            # Add comment with error message
            comment = Comment(f"❌ ERROR:\n{error.message}", "Validation System")
            comment.width = 300
            comment.height = 80
            cell.comment = comment
        else:
            # If column not found, highlight entire row
            self._highlight_entire_row(excel_row, error)
    
    def _highlight_entire_row(self, row_num: int, error):
        """
        Highlight entire row when specific column cannot be determined.
        
        Args:
            row_num: Row number to highlight
            error: ValidationError object
        """
        for col_idx in range(1, min(self.ws.max_column + 1, 15)):  # Limit to reasonable columns
            cell = self.ws.cell(row=row_num, column=col_idx)
            cell.fill = self.ERROR_FILL
        
        # Add comment to first cell
        first_cell = self.ws.cell(row=row_num, column=1)
        comment = Comment(
            f"❌ ERROR in column '{error.column}':\n{error.message}", 
            "Validation System"
        )
        comment.width = 300
        comment.height = 80
        first_cell.comment = comment
    
    def _get_output_path(self) -> str:
        """
        Generate output path for error report.
        
        Returns:
            Path to output file
        """
        # Create temporary file
        base_name = os.path.basename(self.excel_path)
        name_parts = os.path.splitext(base_name)
        error_filename = f"{name_parts[0]}_ERRORS{name_parts[1]}"
        
        # Use temp directory
        output_path = os.path.join(tempfile.gettempdir(), error_filename)
        
        return output_path


def generate_error_excel(excel_path: str, validation_errors: List) -> str:
    """
    Convenience function to generate error report Excel.
    
    Args:
        excel_path: Path to original Excel file
        validation_errors: List of ValidationError objects
    
    Returns:
        Path to generated error report Excel file
    """
    reporter = ErrorReporter(excel_path, validation_errors)
    return reporter.generate_error_report()
