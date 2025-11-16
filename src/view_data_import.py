"""
Events Tracker - View Data Import Module
=========================================
Created: 2025-11-15 17:00 UTC
Last Modified: 2025-11-15 18:30 UTC
Python: 3.11

Description:
Import edited Excel file, detect changes, show diff, and apply updates.
Supports change detection by Event_ID with detailed diff viewer.
"""
import streamlit as st
import pandas as pd
from datetime import datetime
from supabase import Client
from openpyxl import load_workbook
import io


def parse_uploaded_excel(uploaded_file) -> pd.DataFrame:
    """Parse uploaded Excel file and return DataFrame"""
    
    try:
        # Load workbook
        wb = load_workbook(uploaded_file, data_only=True)
        
        # Get Events sheet
        if "Events" not in wb.sheetnames:
            st.error("❌ Excel file must contain 'Events' sheet")
            return None
        
        ws = wb["Events"]
        
        # Read data
        data = []
        headers = []
        
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx == 1:
                # Header row
                headers = [str(cell) if cell is not None else f"Column_{i}" for i, cell in enumerate(row)]
            else:
                # Data rows
                if any(cell is not None for cell in row):  # Skip empty rows
                    data.append(row)
        
        if not data:
            st.error("❌ No data found in Excel file")
            return None
        
        # Create DataFrame
        df = pd.DataFrame(data, columns=headers)
        
        # Validate required columns
        required_cols = ["Event_ID", "Category_Path", "Date"]
        missing_cols = [col for col in required_cols if col not in df.columns]
        
        if missing_cols:
            st.error(f"❌ Missing required columns: {', '.join(missing_cols)}")
            return None
        
        return df
    
    except Exception as e:
        st.error(f"❌ Error parsing Excel file: {str(e)}")
        return None


def get_current_event_data(client: Client, user_id: str, event_id: str) -> dict:
    """Get current event data from database"""
    
    # Get event
    event_response = client.table("events").select(
        "id, category_id, event_date, comment, categories(name, area_id, areas(name))"
    ).eq("id", event_id).eq("user_id", user_id).execute()
    
    if not event_response.data:
        return None
    
    event = event_response.data[0]
    
    # Get event_attributes
    event_data_response = client.table("event_attributes").select(
        "attribute_id, value_numeric, value_text, attribute_definitions(name, data_type)"
    ).eq("event_id", event_id).execute()
    
    event_data = event_data_response.data if event_data_response.data else []
    
    # Build attribute dict
    attributes = {}
    for ed in event_data:
        attr_name = ed["attribute_definitions"]["name"]
        data_type = ed["attribute_definitions"]["data_type"]
        
        if data_type == "numeric":
            attributes[attr_name] = ed.get("value_numeric")
        else:
            attributes[attr_name] = ed.get("value_text")
    
    return {
        "id": event["id"],
        "category_id": event["category_id"],
        "event_date": event["event_date"],
        "comment": event.get("comment"),
        "attributes": attributes,
        "category_path": f"{event['categories']['areas']['name']} > {event['categories']['name']}" if event.get('categories') else "Unknown"
    }


def detect_changes(client: Client, user_id: str, df: pd.DataFrame) -> dict:
    """Detect changes between uploaded data and database"""
    
    changes = {
        "modified": [],
        "errors": []
    }
    
    # Get all attribute columns (exclude fixed columns)
    fixed_columns = ["Event_ID", "Category_Path", "Date", "Comment"]
    attribute_columns = [col for col in df.columns if col not in fixed_columns]
    
    # Process each row
    for idx, row in df.iterrows():
        event_id = row["Event_ID"]
        
        if pd.isna(event_id) or not event_id:
            changes["errors"].append(f"Row {idx + 2}: Missing Event_ID")
            continue
        
        # Get current data from database
        current_data = get_current_event_data(client, user_id, str(event_id))
        
        if not current_data:
            changes["errors"].append(f"Row {idx + 2}: Event_ID {event_id} not found in database")
            continue
        
        # Compare values
        row_changes = []
        
        # Check attributes
        for attr_name in attribute_columns:
            new_value = row[attr_name]
            old_value = current_data["attributes"].get(attr_name)
            
            # Normalize values for comparison
            if pd.isna(new_value):
                new_value = None
            elif isinstance(new_value, str):
                new_value = new_value.strip()
                if new_value == "":
                    new_value = None
            
            if pd.isna(old_value):
                old_value = None
            
            # Compare
            if new_value != old_value:
                row_changes.append({
                    "field": attr_name,
                    "old_value": old_value,
                    "new_value": new_value
                })
        
        # Check comment
        new_comment = row.get("Comment")
        old_comment = current_data.get("comment")
        
        if pd.isna(new_comment):
            new_comment = None
        elif isinstance(new_comment, str):
            new_comment = new_comment.strip()
            if new_comment == "":
                new_comment = None
        
        if pd.isna(old_comment):
            old_comment = None
        
        if new_comment != old_comment:
            row_changes.append({
                "field": "Comment",
                "old_value": old_comment,
                "new_value": new_comment
            })
        
        # If changes detected, add to list
        if row_changes:
            changes["modified"].append({
                "event_id": event_id,
                "category_path": current_data["category_path"],
                "date": current_data["event_date"],
                "changes": row_changes
            })
    
    return changes


def apply_changes(client: Client, user_id: str, modified_changes: list[dict]) -> tuple[int, int, list[str]]:
    """Apply changes to database"""
    
    success_count = 0
    fail_count = 0
    errors = []
    
    for change in modified_changes:
        event_id = change["event_id"]
        
        try:
            # Get event to verify ownership
            event_response = client.table("events").select("id, category_id").eq("id", event_id).eq("user_id", user_id).execute()
            
            if not event_response.data:
                fail_count += 1
                errors.append(f"Event {event_id}: Not found or access denied")
                continue
            
            event = event_response.data[0]
            category_id = event["category_id"]
            
            # Get attributes for this category
            attributes_response = client.table("attribute_definitions").select("id, name, data_type").eq("category_id", category_id).execute()
            attributes = {attr["name"]: attr for attr in attributes_response.data} if attributes_response.data else {}
            
            # Process each change
            for field_change in change["changes"]:
                field = field_change["field"]
                new_value = field_change["new_value"]
                
                if field == "Comment":
                    # Update comment in events table
                    client.table("events").update({
                        "comment": new_value
                    }).eq("id", event_id).execute()
                
                else:
                    # Update attribute value
                    if field not in attributes:
                        errors.append(f"Event {event_id}: Attribute '{field}' not found in category")
                        continue
                    
                    attr = attributes[field]
                    attr_id = attr["id"]
                    data_type = attr["data_type"]
                    
                    # Check if event_attributes exists
                    existing_response = client.table("event_attributes").select("id").eq("event_id", event_id).eq("attribute_id", attr_id).execute()
                    
                    if new_value is None:
                        # Delete if exists
                        if existing_response.data:
                            client.table("event_attributes").delete().eq("event_id", event_id).eq("attribute_id", attr_id).execute()
                    else:
                        # Prepare update data
                        if data_type == "numeric":
                            try:
                                numeric_value = float(new_value)
                                update_data = {
                                    "value_numeric": numeric_value,
                                    "value_text": None
                                }
                            except (ValueError, TypeError):
                                errors.append(f"Event {event_id}: Invalid numeric value for '{field}': {new_value}")
                                continue
                        else:
                            update_data = {
                                "value_numeric": None,
                                "value_text": str(new_value)
                            }
                        
                        if existing_response.data:
                            # Update existing
                            client.table("event_attributes").update(update_data).eq("event_id", event_id).eq("attribute_id", attr_id).execute()
                        else:
                            # Insert new
                            update_data["event_id"] = event_id
                            update_data["attribute_id"] = attr_id
                            client.table("event_attributes").insert(update_data).execute()
            
            success_count += 1
        
        except Exception as e:
            fail_count += 1
            errors.append(f"Event {event_id}: {str(e)}")
    
    return success_count, fail_count, errors


def render_view_data_import(client: Client, user_id: str):
    """Render View Data Import UI"""
    
    st.title("📤 View Data - Import Changes")
    st.markdown("Upload edited Excel file to import changes")
    
    st.info("💡 **Workflow:** Export → Edit in Excel → Upload here → Review changes → Confirm")
    
    st.markdown("---")
    
    # Upload file
    uploaded_file = st.file_uploader(
        "Upload Edited Excel File",
        type=["xlsx"],
        help="Upload the Excel file you downloaded and edited from View Data Export"
    )
    
    if not uploaded_file:
        st.markdown("### 📋 Instructions")
        st.markdown("""
        1. **Export data** using 'View Data - Export' page
        2. **Edit** BLUE columns in Excel (attribute values)
        3. **Save** the Excel file
        4. **Upload** the file here
        5. **Review** detected changes
        6. **Confirm** to apply changes to database
        
        ⚠️ **Important:** Do NOT edit PINK columns (Event_ID, Category_Path, Date)
        """)
        return
    
    # Parse uploaded file
    with st.spinner("Parsing Excel file..."):
        df = parse_uploaded_excel(uploaded_file)
    
    if df is None:
        return
    
    st.success(f"✅ Excel file parsed: {len(df)} events found")
    
    # Preview data
    with st.expander("📄 Preview Uploaded Data", expanded=False):
        st.dataframe(df.head(10), use_container_width=True)
        if len(df) > 10:
            st.caption(f"Showing first 10 of {len(df)} events")
    
    st.markdown("---")
    
    # Detect changes
    with st.spinner("Detecting changes..."):
        changes = detect_changes(client, user_id, df)
    
    modified_changes = changes["modified"]
    errors = changes["errors"]
    
    # Show errors if any
    if errors:
        st.error(f"⚠️ {len(errors)} errors detected:")
        with st.expander("View Errors", expanded=True):
            for error in errors:
                st.write(f"- {error}")
    
    # Show changes
    if not modified_changes:
        if not errors:
            st.success("✅ No changes detected. All values match database.")
        else:
            st.warning("⚠️ No valid changes detected due to errors.")
        return
    
    st.markdown(f"### 📝 Changes Detected: {len(modified_changes)} events")
    
    with st.expander("🔍 View Detailed Changes", expanded=True):
        for change in modified_changes[:10]:  # Show first 10
            st.markdown(f"**Event:** {change['category_path']} | {change['date']}")
            
            changes_df = pd.DataFrame([
                {
                    'Field': c['field'],
                    'Old Value': c['old_value'] if c['old_value'] is not None else '(empty)',
                    'New Value': c['new_value'] if c['new_value'] is not None else '(empty)'
                }
                for c in change['changes']
            ])
            
            st.dataframe(changes_df, use_container_width=True, hide_index=True)
            st.markdown("---")
        
        if len(modified_changes) > 10:
            st.caption(f"... and {len(modified_changes) - 10} more modified events")
    
    st.markdown("---")
    
    # Confirmation
    st.markdown("### ✅ Apply Changes")
    
    st.warning(f"⚠️ This will update {len(modified_changes)} events in the database")
    
    col1, col2, col3 = st.columns([1, 1, 1])
    
    with col2:
        if st.button("🚀 Confirm & Apply Changes", type="primary", use_container_width=True):
            with st.spinner("Applying changes..."):
                success_count, fail_count, apply_errors = apply_changes(client, user_id, modified_changes)
            
            st.markdown("---")
            
            if fail_count == 0:
                st.success(f"🎉 Successfully updated {success_count} events!")
                st.balloons()
            else:
                st.warning(f"⚠️ Updated {success_count} events, {fail_count} failed")
                
                if apply_errors:
                    with st.expander("Error Details"):
                        for error in apply_errors:
                            st.write(f"- {error}")


if __name__ == "__main__":
    st.write("This module should be imported by streamlit_app.py")
