"""
Events Tracker - View Data Export Module
=========================================
Created: 2025-11-15 17:00 UTC
Last Modified: 2025-11-15 18:30 UTC
Python: 3.11

Description:
Export events to Excel for viewing and editing with filters and color-coding.
PINK columns = read-only, BLUE columns = editable.
"""
import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from supabase import Client
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import io


def get_all_areas_categories(client: Client, user_id: str) -> tuple[list[dict], list[dict]]:
    """Get all areas and categories for filters"""
    
    # Get areas
    areas_response = client.table("areas").select("*").eq("user_id", user_id).order("sort_order").execute()
    areas = areas_response.data if areas_response.data else []
    
    # Get categories with full paths
    categories_response = client.table("categories").select("*").eq("user_id", user_id).order("area_id, sort_order").execute()
    categories = categories_response.data if categories_response.data else []
    
    # Build category paths
    for cat in categories:
        area = next((a for a in areas if a["id"] == cat["area_id"]), None)
        if area:
            cat["category_path"] = f"{area['name']} > {cat['name']}"
        else:
            cat["category_path"] = cat["name"]
    
    return areas, categories


def get_category_attributes(client: Client, category_id: str) -> list[dict]:
    """Get all attributes for a category"""
    
    response = client.table("attributes").select("*").eq("category_id", category_id).order("sort_order").execute()
    
    return response.data if response.data else []


def get_events_with_data(
    client: Client,
    user_id: str,
    category_ids: list[str] = None,
    date_from: datetime = None,
    date_to: datetime = None
) -> list[dict]:
    """Get events with their attribute data"""
    
    # Build query
    query = client.table("events").select(
        "id, category_id, event_date, comment, categories(name, area_id, areas(name))"
    ).eq("user_id", user_id)
    
    # Apply filters
    if category_ids:
        query = query.in_("category_id", category_ids)
    
    if date_from:
        query = query.gte("event_date", date_from.strftime("%Y-%m-%d"))
    
    if date_to:
        query = query.lte("event_date", date_to.strftime("%Y-%m-%d"))
    
    query = query.order("event_date", desc=True)
    
    events_response = query.execute()
    events = events_response.data if events_response.data else []
    
    if not events:
        return []
    
    # Get all event IDs
    event_ids = [e["id"] for e in events]
    
    # Get all event_data for these events
    event_data_response = client.table("event_data").select(
        "event_id, attribute_id, value_numeric, value_text, attributes(name, data_type)"
    ).in_("event_id", event_ids).execute()
    
    event_data = event_data_response.data if event_data_response.data else []
    
    # Group event_data by event_id
    data_by_event = {}
    for ed in event_data:
        event_id = ed["event_id"]
        if event_id not in data_by_event:
            data_by_event[event_id] = []
        data_by_event[event_id].append(ed)
    
    # Attach data to events
    for event in events:
        event["event_data"] = data_by_event.get(event["id"], [])
        
        # Build category_path
        if event.get("categories") and event["categories"].get("areas"):
            event["category_path"] = f"{event['categories']['areas']['name']} > {event['categories']['name']}"
        elif event.get("categories"):
            event["category_path"] = event["categories"]["name"]
        else:
            event["category_path"] = "Unknown"
    
    return events


def create_excel_export(events: list[dict], selected_attributes: list[str] = None) -> bytes:
    """Create Excel file with events data"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Events"
    
    # Define colors
    pink_fill = PatternFill(start_color="FFE6F0", end_color="FFE6F0", fill_type="solid")  # Read-only
    blue_fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")  # Editable
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Collect all unique attributes if not specified
    if selected_attributes is None:
        selected_attributes = set()
        for event in events:
            for ed in event.get("event_data", []):
                attr_name = ed["attributes"]["name"]
                selected_attributes.add(attr_name)
        selected_attributes = sorted(list(selected_attributes))
    
    # Define columns
    fixed_columns = ["Event_ID", "Category_Path", "Date"]
    attribute_columns = selected_attributes
    meta_columns = ["Comment"]
    
    all_columns = fixed_columns + attribute_columns + meta_columns
    
    # Write header row
    for col_idx, col_name in enumerate(all_columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    # Write data rows
    for row_idx, event in enumerate(events, start=2):
        # Fixed columns (READ-ONLY - PINK)
        for col_idx, col_name in enumerate(fixed_columns, start=1):
            if col_name == "Event_ID":
                value = event["id"]
            elif col_name == "Category_Path":
                value = event["category_path"]
            elif col_name == "Date":
                value = event["event_date"]
            else:
                value = ""
            
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = pink_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Attribute columns (EDITABLE - BLUE)
        # Build attribute data dict
        attr_data = {}
        for ed in event.get("event_data", []):
            attr_name = ed["attributes"]["name"]
            data_type = ed["attributes"]["data_type"]
            
            if data_type == "numeric":
                attr_data[attr_name] = ed.get("value_numeric")
            else:
                attr_data[attr_name] = ed.get("value_text")
        
        for col_idx, attr_name in enumerate(attribute_columns, start=len(fixed_columns) + 1):
            value = attr_data.get(attr_name, "")
            
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = blue_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="right" if isinstance(value, (int, float)) else "left", vertical="center")
        
        # Meta columns (EDITABLE - BLUE)
        for col_idx, col_name in enumerate(meta_columns, start=len(fixed_columns) + len(attribute_columns) + 1):
            if col_name == "Comment":
                value = event.get("comment", "")
            else:
                value = ""
            
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = blue_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Auto-size columns
    for col_idx in range(1, len(all_columns) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[col_letter]:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Freeze panes (header + first 3 columns)
    ws.freeze_panes = "D2"
    
    # Add Instructions sheet
    ws_instructions = wb.create_sheet("Instructions")
    instructions = [
        ["VIEW DATA EXPORT - Instructions"],
        [""],
        ["WHAT IS THIS FILE?"],
        ["This Excel file contains your events exported from the Events Tracker app."],
        ["You can view and EDIT this data, then re-import it back into the app."],
        [""],
        ["COLOR CODING:"],
        ["PINK columns = READ-ONLY (Event_ID, Category_Path, Date)"],
        ["  - DO NOT edit these columns"],
        ["  - They identify which event to update"],
        [""],
        ["BLUE columns = EDITABLE (Attribute values, Comment)"],
        ["  - You CAN edit these columns"],
        ["  - Changes will be detected and applied on import"],
        [""],
        ["HOW TO EDIT:"],
        ["1. Edit any BLUE column values"],
        ["2. Save this Excel file"],
        ["3. Go to app > 'View Data' page"],
        ["4. Click 'Import Changes'"],
        ["5. Upload this file"],
        ["6. Review detected changes (DIFF view)"],
        ["7. Confirm to apply changes"],
        [""],
        ["IMPORTANT:"],
        ["- Do NOT delete rows (use Delete in app instead)"],
        ["- Do NOT add rows (use Add Event or Bulk Import instead)"],
        ["- Do NOT edit PINK columns"],
        ["- Empty cells mean 'no value' for that attribute"],
        [""],
        ["SAFETY:"],
        ["- Only changed values are updated"],
        ["- Unchanged values remain untouched"],
        ["- Full audit log of changes maintained"],
    ]
    
    for row_idx, row_data in enumerate(instructions, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_instructions.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif value in ["COLOR CODING:", "HOW TO EDIT:", "IMPORTANT:", "SAFETY:", "WHAT IS THIS FILE?"]:
                cell.font = Font(bold=True, size=11)
    
    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()


def render_view_data_export(client: Client, user_id: str):
    """Render View Data Export UI"""
    
    st.title("📊 View Data - Export")
    st.markdown("Export events to Excel for viewing and editing")
    
    # Get areas and categories
    areas, categories = get_all_areas_categories(client, user_id)
    
    if not categories:
        st.warning("⚠️ No categories found. Please create structure first.")
        return
    
    st.markdown("---")
    
    # Filters
    st.markdown("### 🔍 Filter Events")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        # Category filter
        category_options = ["All Categories"] + [cat["category_path"] for cat in categories]
        selected_category = st.selectbox("Category", category_options)
        
        if selected_category == "All Categories":
            category_ids = [cat["id"] for cat in categories]
        else:
            selected_cat = next((c for c in categories if c["category_path"] == selected_category), None)
            category_ids = [selected_cat["id"]] if selected_cat else []
    
    with col2:
        # Date from
        default_from = datetime.now() - timedelta(days=30)
        date_from = st.date_input("Date From", value=default_from)
    
    with col3:
        # Date to
        date_to = st.date_input("Date To", value=datetime.now())
    
    st.markdown("---")
    
    # Attribute selection (optional)
    if selected_category != "All Categories" and category_ids:
        selected_cat = next((c for c in categories if c["category_path"] == selected_category), None)
        if selected_cat:
            attributes = get_category_attributes(client, selected_cat["id"])
            
            if attributes:
                st.markdown("### 📋 Select Attributes to Export")
                
                select_all = st.checkbox("Select All Attributes", value=True)
                
                if select_all:
                    selected_attrs = [attr["name"] for attr in attributes]
                else:
                    selected_attrs = st.multiselect(
                        "Attributes",
                        options=[attr["name"] for attr in attributes],
                        default=[attr["name"] for attr in attributes]
                    )
            else:
                selected_attrs = None
        else:
            selected_attrs = None
    else:
        selected_attrs = None  # All attributes for all categories
    
    st.markdown("---")
    
    # Export button
    col1, col2, col3 = st.columns([1, 1, 1])
    
    with col2:
        if st.button("📥 Export to Excel", type="primary", use_container_width=True):
            with st.spinner("Fetching events..."):
                events = get_events_with_data(
                    client=client,
                    user_id=user_id,
                    category_ids=category_ids if category_ids else None,
                    date_from=datetime.combine(date_from, datetime.min.time()) if date_from else None,
                    date_to=datetime.combine(date_to, datetime.max.time()) if date_to else None
                )
            
            if not events:
                st.warning("⚠️ No events found with selected filters.")
            else:
                with st.spinner(f"Creating Excel file with {len(events)} events..."):
                    excel_bytes = create_excel_export(events, selected_attrs)
                
                st.success(f"✅ Excel file created with {len(events)} events!")
                
                # Download button
                filename = f"events_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                
                st.download_button(
                    label="📥 Download Excel File",
                    data=excel_bytes,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
                st.info("💡 **Next Steps:**\n1. Download the file\n2. Edit BLUE columns in Excel\n3. Save the file\n4. Go to 'Import Changes' to upload edited file")


if __name__ == "__main__":
    st.write("This module should be imported by streamlit_app.py")
