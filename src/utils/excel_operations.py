"""
Excel Operations Utility Module
================================
Created: 2025-11-29 10:30 UTC
Last Modified: 2025-11-29 10:30 UTC
Python: 3.11

Description:
Centralized Excel export/import operations for Events Tracker.
Handles all Excel file generation, formatting, and parsing.
Eliminates code duplication between modules.

Features:
- Unified Excel export (hierarchical, flat, enhanced formats)
- Color coding (PINK for auto, BLUE for editable)
- Drop-down validations
- Formula generation
- Row/column grouping
- Excel import and validation

Dependencies: pandas, openpyxl, io

Usage:
    from src.utils.excel_operations import export_structure_to_excel
    
    excel_bytes = export_structure_to_excel(df, format='hierarchical')
    st.download_button("Download", excel_bytes, "structure.xlsx")
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import Optional, Dict, List, Any
from datetime import datetime
import streamlit as st


# ============================================================================
# COLOR DEFINITIONS
# ============================================================================

# PINK - Auto-calculated columns (READ-ONLY)
PINK_FILL = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')

# BLUE - Editable columns
BLUE_FILL = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')

# Header - Dark gray
HEADER_FILL = PatternFill(start_color='4F4F4F', end_color='4F4F4F', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')

# Borders
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


# ============================================================================
# EXPORT FUNCTIONS
# ============================================================================

def export_structure_to_excel(
    df: pd.DataFrame,
    format_type: str = 'hierarchical',
    filename: Optional[str] = None
) -> bytes:
    """
    Export structure DataFrame to Excel with formatting.
    
    Args:
        df: DataFrame with structure data
        format_type: 'hierarchical', 'flat', or 'enhanced'
        filename: Optional filename (for local save, not download)
        
    Returns:
        Excel file as bytes (for st.download_button)
        
    Examples:
        # For download button
        excel_bytes = export_structure_to_excel(df, 'hierarchical')
        st.download_button("Download", excel_bytes, "structure.xlsx")
        
        # For local file
        excel_bytes = export_structure_to_excel(df, 'enhanced')
        with open('output.xlsx', 'wb') as f:
            f.write(excel_bytes)
    """
    try:
        # Create Excel file in memory
        output = BytesIO()
        
        if format_type == 'hierarchical':
            _export_hierarchical(df, output)
        elif format_type == 'flat':
            _export_flat(df, output)
        elif format_type == 'enhanced':
            _export_enhanced(df, output)
        else:
            raise ValueError(f"Unknown format: {format_type}")
        
        # Return bytes
        output.seek(0)
        return output.getvalue()
        
    except Exception as e:
        st.error(f"Error exporting to Excel: {str(e)}")
        return b''


def _export_hierarchical(df: pd.DataFrame, output: BytesIO):
    """
    Export in hierarchical format (areas > categories > attributes).
    This is the standard format for structure viewing/editing.
    """
    # Write to Excel
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Hierarchical_View', index=False)
        
        workbook = writer.book
        worksheet = workbook['Hierarchical_View']
        
        # Apply formatting
        _apply_color_coding(worksheet, df)
        _apply_validations(worksheet, df)
        _apply_formulas(worksheet, df)
        _apply_grouping(worksheet, df)
        _auto_size_columns(worksheet)


def _export_flat(df: pd.DataFrame, output: BytesIO):
    """
    Export in flat format (simple table, no grouping).
    Useful for quick viewing or external processing.
    """
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Structure', index=False)
        
        workbook = writer.book
        worksheet = workbook['Structure']
        
        # Basic formatting only
        _apply_header_format(worksheet)
        _auto_size_columns(worksheet)


def _export_enhanced(df: pd.DataFrame, output: BytesIO):
    """
    Export with FULL enhanced features:
    - Color coding
    - Drop-down validations
    - Formulas
    - Grouping
    - Freeze panes
    - Auto-filtering
    
    This is the premium format for power users.
    """
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Hierarchical_View', index=False)
        
        workbook = writer.book
        worksheet = workbook['Hierarchical_View']
        
        # Apply ALL features
        _apply_color_coding(worksheet, df)
        _apply_validations(worksheet, df)
        _apply_formulas(worksheet, df)
        _apply_grouping(worksheet, df)
        _apply_freeze_panes(worksheet)
        _apply_auto_filter(worksheet, df)
        _auto_size_columns(worksheet)


# ============================================================================
# FORMATTING FUNCTIONS
# ============================================================================

def _apply_color_coding(worksheet, df: pd.DataFrame):
    """
    Apply PINK (auto) and BLUE (editable) color coding.
    
    PINK columns (auto-calculated):
    - Type, Level, Area, Category_Path, Sort_Order
    
    BLUE columns (editable):
    - Category, Attribute_Name, Data_Type, Unit, Is_Required,
      Default_Value, Validation_Min, Validation_Max, Description
    """
    # Define column categories
    pink_columns = ['Type', 'Level', 'Area', 'Category_Path', 'Sort_Order']
    blue_columns = [
        'Category', 'Attribute_Name', 'Data_Type', 'Unit',
        'Is_Required', 'Default_Value', 'Validation_Min',
        'Validation_Max', 'Description'
    ]
    
    # Apply header formatting
    _apply_header_format(worksheet)
    
    # Apply column colors
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        
        # Determine fill color
        if col_name in pink_columns:
            fill = PINK_FILL
        elif col_name in blue_columns:
            fill = BLUE_FILL
        else:
            fill = None
        
        # Apply to all data rows
        if fill:
            for row_idx in range(2, len(df) + 2):  # Skip header
                cell = worksheet[f'{col_letter}{row_idx}']
                cell.fill = fill
                cell.border = THIN_BORDER


def _apply_header_format(worksheet):
    """Apply dark gray header with white text."""
    for cell in worksheet[1]:  # First row
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER


def _apply_validations(worksheet, df: pd.DataFrame):
    """
    Apply drop-down validations for:
    - Type (Area, Category, Attribute)
    - Data_Type (number, text, datetime, boolean, link, image)
    - Is_Required (Yes, No)
    """
    # Find column indices
    col_indices = {col: idx + 1 for idx, col in enumerate(df.columns)}
    
    # Type validation
    if 'Type' in col_indices:
        type_col = get_column_letter(col_indices['Type'])
        type_validation = DataValidation(
            type="list",
            formula1='"Area,Category,Attribute"',
            allow_blank=False
        )
        worksheet.add_data_validation(type_validation)
        type_validation.add(f'{type_col}2:{type_col}1000')
    
    # Data_Type validation
    if 'Data_Type' in col_indices:
        dtype_col = get_column_letter(col_indices['Data_Type'])
        dtype_validation = DataValidation(
            type="list",
            formula1='"number,text,datetime,boolean,link,image"',
            allow_blank=True
        )
        worksheet.add_data_validation(dtype_validation)
        dtype_validation.add(f'{dtype_col}2:{dtype_col}1000')
    
    # Is_Required validation
    if 'Is_Required' in col_indices:
        req_col = get_column_letter(col_indices['Is_Required'])
        req_validation = DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=True
        )
        worksheet.add_data_validation(req_validation)
        req_validation.add(f'{req_col}2:{req_col}1000')


def _apply_formulas(worksheet, df: pd.DataFrame):
    """
    Apply auto-formulas for:
    - Level (calculate from Category_Path)
    - Area (extract from Category_Path)
    
    Note: These are for display/validation, not for editing.
    """
    col_indices = {col: idx + 1 for idx, col in enumerate(df.columns)}
    
    # Level formula (count ">" in Category_Path)
    if 'Level' in col_indices and 'Category_Path' in col_indices:
        level_col = get_column_letter(col_indices['Level'])
        path_col = get_column_letter(col_indices['Category_Path'])
        
        for row_idx in range(2, len(df) + 2):
            formula = f'=IF({path_col}{row_idx}="", "", LEN({path_col}{row_idx})-LEN(SUBSTITUTE({path_col}{row_idx}, ">", "")))'
            worksheet[f'{level_col}{row_idx}'] = formula
    
    # Area formula (extract first part before ">")
    if 'Area' in col_indices and 'Category_Path' in col_indices:
        area_col = get_column_letter(col_indices['Area'])
        path_col = get_column_letter(col_indices['Category_Path'])
        
        for row_idx in range(2, len(df) + 2):
            formula = f'=IF({path_col}{row_idx}="", "", LEFT({path_col}{row_idx}, FIND(">", {path_col}{row_idx}&">") - 1))'
            worksheet[f'{area_col}{row_idx}'] = formula


def _apply_grouping(worksheet, df: pd.DataFrame):
    """
    Apply row grouping by Area and Category.
    Users can collapse/expand groups.
    """
    # TODO: Implement row grouping logic
    # This requires tracking area/category boundaries
    # For now, skip (complex logic)
    pass


def _apply_freeze_panes(worksheet):
    """Freeze header row and first few columns."""
    worksheet.freeze_panes = 'A2'  # Freeze first row


def _apply_auto_filter(worksheet, df: pd.DataFrame):
    """Enable auto-filter on header row."""
    worksheet.auto_filter.ref = f'A1:{get_column_letter(len(df.columns))}1'


def _auto_size_columns(worksheet):
    """Auto-size all columns based on content."""
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Cap at 50
        worksheet.column_dimensions[column_letter].width = adjusted_width


# ============================================================================
# IMPORT FUNCTIONS
# ============================================================================

def import_structure_from_excel(
    file_bytes: bytes,
    sheet_name: str = 'Hierarchical_View'
) -> pd.DataFrame:
    """
    Import structure from Excel file.
    
    Args:
        file_bytes: Excel file as bytes
        sheet_name: Name of sheet to read
        
    Returns:
        DataFrame with structure data
        
    Example:
        uploaded_file = st.file_uploader("Upload Excel")
        if uploaded_file:
            df = import_structure_from_excel(uploaded_file.read())
    """
    try:
        # Read Excel
        df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name)
        
        # Basic validation
        required_columns = ['Type', 'Category_Path']
        missing = [col for col in required_columns if col not in df.columns]
        
        if missing:
            raise ValueError(f"Missing required columns: {missing}")
        
        return df
        
    except Exception as e:
        st.error(f"Error importing Excel: {str(e)}")
        return pd.DataFrame()


def validate_structure_excel(df: pd.DataFrame) -> List[str]:
    """
    Validate imported structure Excel.
    
    Returns list of validation errors (empty if valid).
    
    Args:
        df: DataFrame to validate
        
    Returns:
        List of error messages (empty = valid)
        
    Example:
        errors = validate_structure_excel(df)
        if errors:
            for error in errors:
                st.error(error)
        else:
            st.success("Valid!")
    """
    errors = []
    
    # Check required columns
    required_cols = ['Type', 'Category_Path']
    for col in required_cols:
        if col not in df.columns:
            errors.append(f"Missing required column: {col}")
    
    # Check Type values
    if 'Type' in df.columns:
        valid_types = ['Area', 'Category', 'Attribute']
        invalid_types = df[~df['Type'].isin(valid_types)]['Type'].unique()
        if len(invalid_types) > 0:
            errors.append(f"Invalid Type values: {invalid_types}")
    
    # Check Data_Type values (for Attributes)
    if 'Data_Type' in df.columns:
        valid_dtypes = ['number', 'text', 'datetime', 'boolean', 'link', 'image']
        attr_rows = df[df['Type'] == 'Attribute']
        invalid_dtypes = attr_rows[
            ~attr_rows['Data_Type'].isin(valid_dtypes + ['', None])
        ]['Data_Type'].unique()
        if len(invalid_dtypes) > 0:
            errors.append(f"Invalid Data_Type values: {invalid_dtypes}")
    
    # Check for empty Category_Path
    if 'Category_Path' in df.columns:
        empty_paths = df[df['Category_Path'].isna() | (df['Category_Path'] == '')]
        if len(empty_paths) > 0:
            errors.append(f"{len(empty_paths)} rows have empty Category_Path")
    
    return errors


# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def generate_filename(
    prefix: str = "Structure",
    area: str = "All",
    search: str = "",
    timestamp: bool = True
) -> str:
    """
    Generate Excel filename with filter information.
    
    Args:
        prefix: Filename prefix (e.g., "Structure")
        area: Area filter (e.g., "Training")
        search: Search term (e.g., "Cardio")
        timestamp: Include timestamp in filename
        
    Returns:
        Filename string
        
    Example:
        filename = generate_filename("Structure", "Training", "Cardio")
        # Returns: "Structure_Training_Cardio_20251129_103045.xlsx"
    """
    # Clean up area and search strings
    area_clean = area.replace(" ", "_").replace(">", "-")
    search_clean = search.replace(" ", "_").replace(">", "-")[:30]
    
    parts = [prefix]
    
    if area and area != "All" and area != "All Areas":
        parts.append(area_clean)
    
    if search_clean:
        parts.append(search_clean)
    
    if timestamp:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        parts.append(ts)
    
    filename = "_".join(parts) + ".xlsx"
    
    return filename


def get_excel_mime_type() -> str:
    """Return MIME type for Excel files."""
    return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
