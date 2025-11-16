"""
Events Tracker - Excel Template Validator Module
=================================================
Created: 2025-11-07 13:11 UTC
Last Modified: 2025-11-15 18:30 UTC
Python: 3.11

Description:
Validates Excel templates using names as identifiers.
Detects uniqueness violations within hierarchical scopes.
Highlights errors and provides detailed validation reports.
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.comments import Comment
from typing import Dict, List, Tuple, Set, Optional
import pandas as pd


class ValidationError:
    """Represents a validation error with location and context"""
    
    def __init__(
        self, 
        sheet: str, 
        cell: str, 
        value: str, 
        error_msg: str,
        conflicts: List[str] = None
    ):
        self.sheet = sheet
        self.cell = cell
        self.value = value
        self.error_msg = error_msg
        self.conflicts = conflicts or []
    
    def __str__(self):
        conflict_str = f" (conflicts: {', '.join(self.conflicts)})" if self.conflicts else ""
        return f"[{self.sheet}!{self.cell}] {self.error_msg}: '{self.value}'{conflict_str}"


class HierarchicalValidator:
    """Validates uniqueness within hierarchical scopes"""
    
    def __init__(self):
        # Store {scope_key: {normalized_name: [cell_refs]}}
        self.scope_registry: Dict[str, Dict[str, List[str]]] = {}
        self.errors: List[ValidationError] = []
        
    def determine_scope_key(
        self, 
        area: str, 
        parent: Optional[str], 
        level: int,
        object_type: str
    ) -> str:
        """
        Generate scope key for uniqueness checking.
        
        Rules:
        - Areas: Must be unique globally (per user)
        - Categories at level 1: Must be unique within area
        - Categories at level > 1: Must be unique within parent
        - Attributes: Must be unique within category
        """
        if object_type == 'area':
            return "area:GLOBAL"
        elif object_type == 'category':
            if level == 1:
                return f"category:area:{area}:level:1"
            else:
                return f"category:area:{area}:parent:{parent}:level:{level}"
        elif object_type == 'attribute':
            return f"attribute:category:{area}"  # area is actually category name here
        
        return f"{object_type}:unknown"
    
    def validate_name(
        self, 
        name: str, 
        scope_key: str, 
        cell_ref: str,
        sheet: str
    ) -> Optional[ValidationError]:
        """
        Check if name is unique within scope.
        
        Returns:
            ValidationError if duplicate found, None otherwise
        """
        normalized = name.strip().lower()
        
        if not normalized:
            return ValidationError(
                sheet, cell_ref, name,
                "Name cannot be empty"
            )
        
        if scope_key not in self.scope_registry:
            self.scope_registry[scope_key] = {}
        
        if normalized in self.scope_registry[scope_key]:
            conflicts = self.scope_registry[scope_key][normalized]
            error = ValidationError(
                sheet, cell_ref, name,
                f"Duplicate name in scope",
                conflicts
            )
            self.errors.append(error)
            return error
        
        # Register this name
        self.scope_registry[scope_key][normalized] = [cell_ref]
        return None


class ExcelTemplateValidator:
    """
    Validates Excel templates with name-based references.
    Highlights errors in yellow with descriptive comments.
    """
    
    def __init__(self, filepath: str):
        self.filepath = filepath
        self.wb = load_workbook(filepath)
        self.validator = HierarchicalValidator()
        
        # Styling for errors
        self.error_fill = PatternFill(
            start_color="FFFF00",  # Yellow
            end_color="FFFF00",
            fill_type="solid"
        )
        self.warning_fill = PatternFill(
            start_color="FFA500",  # Orange
            end_color="FFA500",
            fill_type="solid"
        )
        self.error_font = Font(color="FF0000", bold=True)
        
        self.warnings: List[str] = []
    
    def validate_all(self) -> Tuple[bool, List[ValidationError], List[str]]:
        """
        Run all validations.
        
        Returns:
            (is_valid, errors, warnings)
        """
        # Check required sheets exist
        required_sheets = ['Areas', 'Categories', 'Attributes']
        missing_sheets = [s for s in required_sheets if s not in self.wb.sheetnames]
        
        if missing_sheets:
            error = ValidationError(
                'Workbook', 'N/A', '',
                f"Missing required sheets: {', '.join(missing_sheets)}"
            )
            return False, [error], []
        
        # Validate each sheet
        areas_valid = self.validate_areas_sheet()
        categories_valid = self.validate_categories_sheet()
        attrs_valid = self.validate_attributes_sheet()
        
        # Cross-sheet validation
        self.validate_cross_references()
        
        is_valid = (areas_valid and categories_valid and attrs_valid 
                   and len(self.validator.errors) == 0)
        
        return is_valid, self.validator.errors, self.warnings
    
    def validate_areas_sheet(self) -> bool:
        """Validate Areas sheet for duplicate names"""
        ws = self.wb['Areas']
        
        # Expected columns
        expected_cols = ['area_id', 'area_name', 'icon', 'color', 'sort_order', 'description']
        
        # Validate header
        header_row = 1
        for col_idx, col_name in enumerate(expected_cols, start=1):
            cell_value = ws.cell(header_row, col_idx).value
            if cell_value != col_name:
                self.warnings.append(
                    f"Areas sheet: Expected column '{col_name}' at position {col_idx}, "
                    f"found '{cell_value}'"
                )
        
        # Validate data rows
        for row in range(2, ws.max_row + 1):  # Skip header
            # Column B: area_name
            name_cell_ref = f"B{row}"
            name = ws[name_cell_ref].value
            
            if not name:
                continue  # Skip empty rows
            
            # Validate uniqueness
            scope_key = self.validator.determine_scope_key(
                area="", parent=None, level=1, object_type='area'
            )
            error = self.validator.validate_name(
                name, scope_key, name_cell_ref, 'Areas'
            )
            
            if error:
                self._mark_error(ws, name_cell_ref, error)
        
        return len(self.validator.errors) == 0
    
    def validate_categories_sheet(self) -> bool:
        """Validate Categories with hierarchical scope rules"""
        ws = self.wb['Categories']
        
        # Expected columns
        expected_cols = [
            'category_id', 'area_name', 'parent_category', 
            'category_name', 'level', 'sort_order', 'description'
        ]
        
        for row in range(2, ws.max_row + 1):
            # Read row data
            cat_id = ws[f"A{row}"].value  # category_id (optional)
            area_name = ws[f"B{row}"].value  # area_name (required)
            parent_name = ws[f"C{row}"].value  # parent_category (optional)
            cat_name = ws[f"D{row}"].value  # category_name (required)
            level = ws[f"E{row}"].value  # level (required)
            
            if not cat_name or not area_name:
                continue  # Skip incomplete rows
            
            # Validate level is numeric
            if not isinstance(level, (int, float)):
                error = ValidationError(
                    'Categories', f"E{row}", str(level),
                    "Level must be a number"
                )
                self.validator.errors.append(error)
                self._mark_error(ws, f"E{row}", error)
                continue
            
            level = int(level)
            
            # Validate level constraints
            if level < 1 or level > 10:
                error = ValidationError(
                    'Categories', f"E{row}", str(level),
                    "Level must be between 1 and 10"
                )
                self.validator.errors.append(error)
                self._mark_error(ws, f"E{row}", error)
                continue
            
            # Validate parent consistency
            if level == 1 and parent_name:
                self.warnings.append(
                    f"Categories row {row}: Level 1 category '{cat_name}' "
                    f"should not have a parent"
                )
            elif level > 1 and not parent_name:
                error = ValidationError(
                    'Categories', f"C{row}", str(parent_name or ''),
                    f"Level {level} category requires a parent"
                )
                self.validator.errors.append(error)
                self._mark_error(ws, f"C{row}", error)
                continue
            
            # Validate name uniqueness within scope
            scope_key = self.validator.determine_scope_key(
                area=area_name, 
                parent=parent_name, 
                level=level,
                object_type='category'
            )
            
            error = self.validator.validate_name(
                cat_name, scope_key, f"D{row}", 'Categories'
            )
            
            if error:
                self._mark_error(ws, f"D{row}", error)
        
        return len([e for e in self.validator.errors if e.sheet == 'Categories']) == 0
    
    def validate_attributes_sheet(self) -> bool:
        """Validate attribute names within category scope"""
        ws = self.wb['Attributes']
        
        valid_data_types = ['number', 'text', 'datetime', 'boolean', 'link', 'image']
        
        for row in range(2, ws.max_row + 1):
            attr_id = ws[f"A{row}"].value  # attribute_id (optional)
            category_name = ws[f"B{row}"].value  # category_name (required)
            attr_name = ws[f"C{row}"].value  # attribute_name (required)
            data_type = ws[f"D{row}"].value  # data_type (required)
            
            if not category_name or not attr_name:
                continue  # Skip incomplete rows
            
            # Validate data_type
            if data_type and data_type not in valid_data_types:
                error = ValidationError(
                    'Attributes', f"D{row}", str(data_type),
                    f"Invalid data type. Must be one of: {', '.join(valid_data_types)}"
                )
                self.validator.errors.append(error)
                self._mark_error(ws, f"D{row}", error)
            
            # Validate name uniqueness within category
            scope_key = self.validator.determine_scope_key(
                area=category_name,  # Using category_name as scope
                parent=None,
                level=1,
                object_type='attribute'
            )
            
            error = self.validator.validate_name(
                attr_name, scope_key, f"C{row}", 'Attributes'
            )
            
            if error:
                self._mark_error(ws, f"C{row}", error)
        
        return len([e for e in self.validator.errors if e.sheet == 'Attributes']) == 0
    
    def validate_cross_references(self):
        """Validate that referenced names exist"""
        # Build registry of existing names
        areas = set()
        categories_by_area = {}
        
        # Read Areas
        ws_areas = self.wb['Areas']
        for row in range(2, ws_areas.max_row + 1):
            area_name = ws_areas[f"B{row}"].value
            if area_name:
                areas.add(area_name.strip())
        
        # Read Categories and build lookup
        ws_categories = self.wb['Categories']
        for row in range(2, ws_categories.max_row + 1):
            area_name = ws_categories[f"B{row}"].value
            cat_name = ws_categories[f"D{row}"].value
            
            if area_name and cat_name:
                area_name = area_name.strip()
                cat_name = cat_name.strip()
                
                # Check area exists
                if area_name not in areas:
                    error = ValidationError(
                        'Categories', f"B{row}", area_name,
                        f"Referenced area '{area_name}' does not exist in Areas sheet"
                    )
                    self.validator.errors.append(error)
                    self._mark_error(ws_categories, f"B{row}", error)
                
                # Build category registry
                if area_name not in categories_by_area:
                    categories_by_area[area_name] = set()
                categories_by_area[area_name].add(cat_name)
        
        # Validate Attributes reference existing categories
        ws_attributes = self.wb['Attributes']
        for row in range(2, ws_attributes.max_row + 1):
            category_name = ws_attributes[f"B{row}"].value
            
            if category_name:
                category_name = category_name.strip()
                
                # Check if category exists in any area
                found = False
                for area_cats in categories_by_area.values():
                    if category_name in area_cats:
                        found = True
                        break
                
                if not found:
                    error = ValidationError(
                        'Attributes', f"B{row}", category_name,
                        f"Referenced category '{category_name}' does not exist"
                    )
                    self.validator.errors.append(error)
                    self._mark_error(ws_attributes, f"B{row}", error)
    
    def _mark_error(self, worksheet, cell_ref: str, error: ValidationError):
        """Highlight cell in yellow and add comment"""
        cell = worksheet[cell_ref]
        
        # Apply yellow background
        cell.fill = self.error_fill
        
        # Build detailed comment message
        comment_text = f"❌ VALIDATION ERROR\n\n{error.error_msg}"
        if error.conflicts:
            comment_text += f"\n\nConflicts with cells: {', '.join(error.conflicts)}"
        
        # Add or update comment
        cell.comment = Comment(comment_text, "Validator")
        cell.comment.width = 300
        cell.comment.height = 100
    
    def save_highlighted(self, output_path: str):
        """Save Excel file with error highlighting"""
        self.wb.save(output_path)
        self.wb.close()
    
    def generate_error_report(self) -> str:
        """Generate text summary of errors"""
        if not self.validator.errors:
            return "✅ No validation errors found"
        
        lines = [f"❌ Found {len(self.validator.errors)} validation error(s):\n"]
        
        for i, error in enumerate(self.validator.errors, 1):
            lines.append(f"{i}. {error}")
        
        if self.warnings:
            lines.append(f"\n⚠️  {len(self.warnings)} warning(s):")
            for warn in self.warnings:
                lines.append(f"  - {warn}")
        
        return "\n".join(lines)


def validate_template(filepath: str) -> Tuple[bool, str, Optional[str]]:
    """
    Validate Excel template and return results.
    
    Returns:
        (is_valid, report, highlighted_file_path)
    """
    validator = ExcelTemplateValidator(filepath)
    is_valid, errors, warnings = validator.validate_all()
    
    report = validator.generate_error_report()
    
    if not is_valid:
        # Save highlighted version
        output_path = filepath.replace('.xlsx', '_errors.xlsx')
        validator.save_highlighted(output_path)
        return False, report, output_path
    
    # Close workbook if no errors
    validator.wb.close()
    return True, report, None
